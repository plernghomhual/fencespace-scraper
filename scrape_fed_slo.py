"""
scrape_fed_slo.py - Sabljaška zveza Slovenije national rankings scraper.

Probe evidence, 2026-06-02:
  - Requested veza.si host did not appear to be the active public federation host.
  - Public rankings page: https://www.sabljaska-zveza.si/rang-lestvice.html
  - Request method: GET
  - Current link format: Google Sheets export linked from "Aktualne rang lestvice".
  - Durable fallback format: per-weapon public PDFs under /uploads/.../rl_24_25_<weapon>.pdf.
  - Public PDF coverage: Senior (Člani) and Junior (Mladinci) Men/Women for
    Foil (Floret), Epee (Meč), and Sabre (Sablja).
"""

from __future__ import annotations

import io
import re
import time
import unicodedata
from datetime import UTC, datetime, timezone

import requests
from bs4 import BeautifulSoup

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import get_state, set_state

try:
    from season_utils import normalize_season, season_to_string
except ImportError:  # pragma: no cover - compatibility fallback
    def season_to_string(season_int: int) -> str:
        return f"{season_int - 1:04d}-{season_int:04d}"

    def normalize_season(raw) -> str:
        if isinstance(raw, int):
            return season_to_string(raw)
        return str(raw)


SOURCE = "slo_fencing"
COUNTRY = "Slovenia"
BASE_URL = "https://www.sabljaska-zveza.si/rang-lestvice.html"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "sl-SI,sl;q=0.9,en;q=0.8",
    "Referer": BASE_URL,
}

CURRENT_SHEET_ID = "1zcnIsSg1a5bTUlGcF89yUMOTTQyp3y9H0H0VU-DB4Sc"
CURRENT_SHEET_XLSX_URL = (
    f"https://docs.google.com/spreadsheets/d/{CURRENT_SHEET_ID}/export?format=xlsx"
)

PDF_RANKING_URLS = {
    "Foil": "https://www.sabljaska-zveza.si/uploads/1/0/9/1/109197245/rl_24_25_floret.pdf",
    "Epee": "https://www.sabljaska-zveza.si/uploads/1/0/9/1/109197245/rl_24_25_me%C4%8D.pdf",
    "Sabre": "https://www.sabljaska-zveza.si/uploads/1/0/9/1/109197245/rl_24_25_sablja.pdf",
}

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

_PDF_TEXT_CACHE: dict[str, str] = {}
_CURRENT_SHEET_TEXT: str | None | bool = False

_NO_DATA_MARKERS = {
    "ni podatkov",
    "no data",
    "no rankings",
    "no ranking",
    "not found",
}
_SKIP_VALUES = {
    "",
    "dns",
    "dnf",
    "dq",
    "dsq",
    "wd",
    "ret",
    "skupaj",
    "summary",
    "subtotal",
    "total",
    "totals",
    "diskvalificiran",
    "diskvalificirana",
}
_CLUB_STARTERS = {
    "SK",
    "SD",
    "ŠD",
    "MK",
    "PK",
    "CS",
    "SG",
    "PENTASCHERMA",
    "FRIULI",
    "SAN",
}
_WEAPON_SL = {
    "Foil": "floret",
    "Epee": "meč",
    "Sabre": "sablja",
}
_GENDER_SL = {
    "Men": "moški",
    "Women": "ženski",
}
_CATEGORY_SL = {
    "Senior": "člani",
    "Junior": "mladinci",
}


def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _normalize_text(value: str) -> str:
    value = unicodedata.normalize("NFKD", _clean_text(value).lower())
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9]+", " ", value).strip()


def _header_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _normalize_text(value))


def _parse_rank(value: str) -> int | None:
    text = _clean_text(value).strip(".")
    key = _header_key(text)
    if key in _SKIP_VALUES:
        return None
    match = re.match(r"^(\d{1,4})(?:[.)])?$", text)
    if not match:
        return None
    return int(match.group(1))


def _parse_points(value: str) -> float | None:
    text = _clean_text(value)
    if not text:
        return None
    match = re.search(r"[-+]?\d[\d.,]*", text)
    if not match:
        return None
    number = match.group(0).replace(" ", "")
    if "," in number and "." in number:
        if number.rfind(",") > number.rfind("."):
            number = number.replace(".", "").replace(",", ".")
        else:
            number = number.replace(",", "")
    elif "," in number:
        number = number.replace(".", "").replace(",", ".")
    try:
        return float(number)
    except ValueError:
        return None


def _find_header_mapping(cells: list[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(cells):
        key = _header_key(value)
        if key in {"rang", "rank", "mesto", "uvrstitev", "no", "st"}:
            mapping.setdefault("rank", index)
        elif key in {"ime", "tekmovalec", "tekmovalka", "name", "fencer", "priimekinime"}:
            mapping.setdefault("name", index)
        elif key in {"klub", "club", "drustvo", "drustvaklub"}:
            mapping.setdefault("club", index)
        elif key in {"tocke", "points", "totalpoints", "skupaj"}:
            mapping.setdefault("points", index)
    if {"rank", "name", "points"}.issubset(mapping):
        return mapping
    return None


def _row_from_cells(cells: list[str], mapping: dict[str, int]) -> dict | None:
    try:
        rank_text = cells[mapping["rank"]]
        name = _clean_text(cells[mapping["name"]])
        points_text = cells[mapping["points"]]
    except (IndexError, KeyError):
        return None

    rank = _parse_rank(rank_text)
    if rank is None or not name or _header_key(name) in _SKIP_VALUES:
        return None

    points = _parse_points(points_text)
    club = None
    club_index = mapping.get("club")
    if club_index is not None and club_index < len(cells):
        club = _clean_text(cells[club_index]) or None

    return {"rank": rank, "name": name, "club": club, "points": points}


def _looks_like_no_data(text: str) -> bool:
    normalized = _normalize_text(text)
    return any(marker in normalized for marker in _NO_DATA_MARKERS)


def _parse_html_tables(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    results: list[dict] = []
    for table in soup.find_all("table"):
        mapping: dict[str, int] | None = None
        for row in table.find_all("tr"):
            cells = [_clean_text(cell.get_text(" ", strip=True)) for cell in row.find_all(["th", "td"])]
            if not cells:
                continue
            header_mapping = _find_header_mapping(cells)
            if header_mapping:
                mapping = header_mapping
                continue
            if mapping is None and len(cells) >= 4:
                mapping = {"rank": 0, "name": 1, "club": 2, "points": 3}
            if mapping is None:
                continue
            parsed = _row_from_cells(cells, mapping)
            if parsed:
                results.append(parsed)
    return results


def _split_name_club(value: str) -> tuple[str, str | None]:
    tokens = _clean_text(value).split()
    if not tokens:
        return "", None

    normalized_starters = {_normalize_text(token).upper() for token in _CLUB_STARTERS}
    for index in range(1, len(tokens)):
        token = _normalize_text(tokens[index]).upper()
        if token in normalized_starters:
            return " ".join(tokens[:index]), " ".join(tokens[index:])

    for index in range(len(tokens) - 1, 0, -1):
        tail = tokens[index:]
        if 1 <= len(tail) <= 5 and all(part.upper() == part for part in tail):
            return " ".join(tokens[:index]), " ".join(tail)

    return " ".join(tokens), None


def _parse_pdf_line(line: str) -> dict | None:
    text = _clean_text(line)
    match = re.match(r"^(\d{1,4})(?:[.)])?\s+(.+?)\s+((?:18|19|20)\d{2})\s+(.+)$", text)
    if not match:
        return None
    rank = _parse_rank(match.group(1))
    if rank is None:
        return None
    name, club = _split_name_club(match.group(2))
    if not name:
        return None
    points = _parse_points(match.group(4))
    return {"rank": rank, "name": name, "club": club, "points": points}


def _parse_delimited_text(text: str) -> list[dict]:
    results: list[dict] = []
    mapping: dict[str, int] | None = None
    for raw_line in text.splitlines():
        line = _clean_text(raw_line)
        if not line:
            continue
        if "\t" in raw_line:
            cells = [_clean_text(cell) for cell in raw_line.split("\t")]
        elif "|" in raw_line:
            cells = [_clean_text(cell) for cell in raw_line.split("|")]
        else:
            parsed = _parse_pdf_line(line)
            if parsed:
                results.append(parsed)
            continue

        cells = [cell for cell in cells if cell]
        if not cells:
            continue
        header_mapping = _find_header_mapping(cells)
        if header_mapping:
            mapping = header_mapping
            continue
        if mapping is None and len(cells) >= 4:
            mapping = {"rank": 0, "name": 1, "club": 2, "points": 3}
        if mapping:
            parsed = _row_from_cells(cells, mapping)
            if parsed:
                results.append(parsed)
    return results


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Slovenian federation ranking HTML, delimited text, or extracted PDF text."""
    if not html_or_text or _looks_like_no_data(html_or_text):
        return []

    html_results = _parse_html_tables(html_or_text)
    if html_results:
        return html_results

    return _parse_delimited_text(html_or_text)


def _matches_combo(line: str, weapon: str, gender: str, category: str) -> bool:
    normalized = _normalize_text(line)
    return (
        _normalize_text(_WEAPON_SL[weapon]) in normalized
        and _normalize_text(_GENDER_SL[gender]) in normalized
        and _normalize_text(_CATEGORY_SL[category]) in normalized
    )


def _is_section_header(line: str) -> bool:
    normalized = _normalize_text(line)
    has_weapon = any(_normalize_text(value) in normalized for value in _WEAPON_SL.values())
    has_gender = any(_normalize_text(value) in normalized for value in _GENDER_SL.values())
    has_category = any(
        value in normalized
        for value in ["clani", "mladinci", "kadeti", "u14", "veterani"]
    )
    return has_weapon and has_gender and has_category


def _extract_combo_section(text: str, weapon: str, gender: str, category: str) -> str | None:
    lines = [_clean_text(line) for line in text.splitlines() if _clean_text(line)]
    start = None
    for index, line in enumerate(lines):
        if _matches_combo(line, weapon, gender, category):
            start = index
            break
    if start is None:
        return None

    end = len(lines)
    for index in range(start + 1, len(lines)):
        if _is_section_header(lines[index]):
            end = index
            break
    return "\n".join(lines[start:end])


def _looks_blocked_or_js_only(text: str) -> bool:
    normalized = _normalize_text(text)
    markers = [
        "access denied",
        "forbidden",
        "servicelogin",
        "sign in",
        "login",
        "enable javascript",
        "javascript required",
    ]
    return any(marker in normalized for marker in markers)


def _xlsx_to_text(content: bytes) -> str | None:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        return None

    parts: list[str] = []
    for worksheet in workbook.worksheets:
        parts.append(f"# {worksheet.title}")
        for row in worksheet.iter_rows(values_only=True):
            values = [_clean_text(value) for value in row]
            if any(values):
                parts.append("\t".join(values))
    return "\n".join(parts)


def _fetch_current_sheet_text(weapon: str, gender: str, category: str) -> str | None:
    global _CURRENT_SHEET_TEXT
    if _CURRENT_SHEET_TEXT is False:
        try:
            response = federation_request(
                "get",
                CURRENT_SHEET_XLSX_URL,
                headers=HEADERS,
                timeout=30,
                allow_redirects=True,
            )
        except requests.RequestException as exc:
            print(f"    Current Google Sheet request error: {exc}")
            _CURRENT_SHEET_TEXT = None
        else:
            content_type = response.headers.get("content-type", "").lower()
            if response.status_code != 200:
                print(f"    Current Google Sheet HTTP {response.status_code}")
                _CURRENT_SHEET_TEXT = None
            elif "html" in content_type and _looks_blocked_or_js_only(response.text):
                print("    Current Google Sheet appears login-only or JS-rendered")
                _CURRENT_SHEET_TEXT = None
            elif response.content.startswith(b"PK"):
                _CURRENT_SHEET_TEXT = _xlsx_to_text(response.content)
            elif "text" in content_type or "csv" in content_type:
                _CURRENT_SHEET_TEXT = response.text
            else:
                _CURRENT_SHEET_TEXT = None

    if not isinstance(_CURRENT_SHEET_TEXT, str) or not _CURRENT_SHEET_TEXT:
        return None

    section = _extract_combo_section(_CURRENT_SHEET_TEXT, weapon, gender, category)
    if section and parse_rankings_table(section):
        return section
    return None


def _extract_pdf_text(content: bytes) -> str:
    try:
        import pdfplumber
    except ImportError:
        print("    pdfplumber is not installed; cannot parse Slovenia ranking PDF")
        return ""

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)


def _download_pdf_text(url: str) -> str | None:
    if url in _PDF_TEXT_CACHE:
        return _PDF_TEXT_CACHE[url]

    for attempt in range(2):
        try:
            response = federation_request(
                "get",
                url,
                headers=HEADERS,
                timeout=30,
                allow_redirects=True,
            )
        except requests.RequestException as exc:
            print(f"    Request error for {url}: {exc}")
            if attempt == 0:
                time.sleep(REQUEST_DELAY)
                continue
            return None

        if response.status_code == 200:
            content_type = response.headers.get("content-type", "").lower()
            if response.content.startswith(b"%PDF") or "pdf" in content_type:
                text = _extract_pdf_text(response.content)
                if text:
                    _PDF_TEXT_CACHE[url] = text
                    return text
                return None
            if _looks_blocked_or_js_only(response.text):
                print(f"    No scrapeable rankings at {url}: login-only or JS-rendered")
                return None
            return response.text

        if response.status_code in {401, 403, 404}:
            print(f"    No scrapeable rankings at {url}: HTTP {response.status_code}")
            return None

        print(f"    HTTP {response.status_code} for {url}")
        if attempt == 0:
            time.sleep(REQUEST_DELAY)

    return None


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Return content for one combo, or None when the public source is missing/blocked."""
    current_sheet_section = _fetch_current_sheet_text(weapon, gender, category)
    if current_sheet_section:
        return current_sheet_section

    url = PDF_RANKING_URLS.get(weapon)
    if not url:
        print(f"    No scrapeable rankings at {BASE_URL}: missing {weapon} source")
        return None

    text = _download_pdf_text(url)
    if not text:
        return None

    section = _extract_combo_section(text, weapon, gender, category)
    if not section:
        print(f"    No scrapeable rankings at {url}: missing {weapon} {gender} {category}")
        return None
    if not parse_rankings_table(section):
        print(f"    No rows parsed for {weapon} {gender} {category} from {url}")
        return None
    return section


def current_season() -> str:
    """Return the current fencing season as YYYY-YYYY."""
    now = datetime.now(UTC)
    season_end_year = now.year if now.month < 7 else now.year + 1
    return normalize_season(season_to_string(season_end_year))


def _combo_label(combo: tuple[str, str, str]) -> str:
    weapon, gender, category = combo
    return f"{weapon} {gender} {category}"


def _source_url_for(weapon: str) -> str:
    return PDF_RANKING_URLS.get(weapon, BASE_URL)


def main():
    run_log = ScraperRunLogger("scrape_fed_slo").start()
    season = current_season()
    previous_state = get_state(SOURCE, "last_ranking_run") or {}
    if previous_state:
        print(f"Previous Slovenia ranking state found for {previous_state.get('season', 'unknown season')}")

    print(f"Slovenia federation rankings - season {season}")
    total_written = 0
    total_failed = 0
    total_skipped = 0
    working_combos: list[str] = []
    failed_combos: list[str] = []

    try:
        for weapon, gender, category in RANKING_COMBOS:
            label = _combo_label((weapon, gender, category))
            print(f"  {label}...")

            content = fetch_rankings_page(weapon, gender, category)
            if content is None:
                total_failed += 1
                failed_combos.append(label)
                time.sleep(REQUEST_DELAY)
                continue

            parsed = parse_rankings_table(content)
            if not parsed:
                print("    No rows parsed")
                total_failed += 1
                failed_combos.append(label)
                time.sleep(REQUEST_DELAY)
                continue

            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                    metadata={
                        "source_url": _source_url_for(weapon),
                        "ranking_page": BASE_URL,
                        "format": "pdf",
                    },
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            print(f"    Parsed {len(parsed)} rows; written {written}")
            total_written += written
            working_combos.append(label)
            time.sleep(REQUEST_DELAY)

        metadata = {
            "season": season,
            "ranking_page": BASE_URL,
            "working_combos": working_combos,
            "failed_combos": failed_combos,
            "combos_attempted": len(RANKING_COMBOS),
            "combos_working": len(working_combos),
        }
        set_state(SOURCE, "last_ranking_run", metadata)
        run_log.complete(written=total_written, failed=total_failed, skipped=total_skipped, metadata=metadata)
        print(
            f"Done - written={total_written}, failed={total_failed}, "
            f"skipped={total_skipped}, combos={len(working_combos)}/{len(RANKING_COMBOS)}"
        )
        if failed_combos:
            print("Failed combos: " + ", ".join(failed_combos))
    except Exception as exc:
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
