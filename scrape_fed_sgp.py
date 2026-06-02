"""
scrape_fed_sgp.py — Fencing Singapore national rankings scraper.

Probe notes, 2026-06-01:
  - https://fencing.org.sg/* did not resolve.
  - https://www.fencingsingapore.org.sg/ranking-files/ is public and lists
    current 25-26 season ranking files.
  - Current ranking files are public XLSX downloads reached by GET from
    WordPress Download Manager `data-downloadurl` links.
  - Latest probed files:
      Epee  -> 25-26-Ranking-Epee_260524.xlsx
      Foil  -> 25-26-Ranking-Foil_260517.xlsx
      Sabre -> 25-26-Ranking-Sabre_260504.xlsx
  - Response format: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.
  - Public sheets cover all 12 required Senior/Junior Men/Women
    Foil/Epee/Sabre combos. Cadet sheets also exist but are out of scope here.
  - https://my.fencingsingapore.org.sg/showranks is public HTML, but 2025/2026
    form queries returned empty tables during probe; older years returned rows.

Relevant workbook columns:
  # | Fencer | Club/School | ... | Final Ranking Points | Final Rank
"""

from __future__ import annotations

import io
import re
import time
from datetime import datetime, timezone
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import get_state, set_state

SOURCE = "sgp_fencing"
COUNTRY = "SGP"
BASE_URL = "https://www.fencingsingapore.org.sg"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    ),
    "Accept-Language": "en-SG,en-US;q=0.8,en;q=0.6",
}

RANKING_FILES_PAGE = f"{BASE_URL}/ranking-files/"
DOWNLOAD_PAGES = {
    "Epee": f"{BASE_URL}/download/25-26-season-ranking-file-epee-updated-9-nov-2025/",
    "Foil": f"{BASE_URL}/download/25-26-season-ranking-file-foil-updated-9-nov-2025/",
    "Sabre": f"{BASE_URL}/download/25-26-season-ranking-file-sabre-updated-2-nov-2025/",
}

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

_WORKBOOK_CACHE: dict[str, bytes] = {}

_SKIP_ROW_TOKENS = {
    "",
    "dns",
    "dq",
    "dsq",
    "dnf",
    "wd",
    "withdrawn",
    "total",
    "totals",
    "summary",
    "subtotal",
}
_NO_DATA_MARKERS = {
    "no rankings available",
    "no ranking available",
    "no data",
    "please select another ranking file",
}


def _clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip() if value is not None else ""


def _header_key(value: str) -> str:
    text = _clean_text(value).lower()
    if text == "#":
        return "#"
    text = text.replace("&", " and ").replace("/", " ")
    return re.sub(r"[^a-z0-9]+", "", text)


def _parse_rank(value: str) -> int | None:
    text = _clean_text(value).rstrip(".")
    if _header_key(text) in _SKIP_ROW_TOKENS:
        return None
    match = re.fullmatch(r"(\d+)(?:\.0+)?", text)
    return int(match.group(1)) if match else None


def _parse_points(value: str) -> float | None:
    text = _clean_text(value).replace(" ", "").replace("\u00a0", "")
    if not text or _header_key(text) in _SKIP_ROW_TOKENS or text in {"-", "—", "–"}:
        return None

    text = re.sub(r"[^0-9,.\-]", "", text)
    if not text or text in {"-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def _points_header_score(key: str) -> int | None:
    if key in {"finalrankingpoints", "finalpoints"}:
        return 0
    if key in {"totalpoints", "rankingpoints", "points", "pts", "score"}:
        return 1
    if key.endswith("points"):
        blocked = ("localpoints", "basepts", "basepoints", "after", "overalllocal")
        if not any(token in key for token in blocked):
            return 2
    return None


def _detect_columns(cells: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    point_candidates: list[tuple[int, int]] = []

    for idx, cell in enumerate(cells):
        key = _header_key(cell)
        if key in {"#", "rank", "ranking", "place", "position", "placing"} and "rank_col" not in mapping:
            mapping["rank_col"] = idx
        elif key in {"name", "fencer", "athlete", "fullname"} and "name_col" not in mapping:
            mapping["name_col"] = idx
        elif key == "clubschool" and "club_col" not in mapping:
            mapping["club_col"] = idx
        elif key in {"club", "clubname", "clubclubs", "team"} and "club_col" not in mapping:
            mapping["club_col"] = idx
        elif key in {"school", "schoolname"} and "school_col" not in mapping:
            mapping["school_col"] = idx

        score = _points_header_score(key)
        if score is not None:
            point_candidates.append((score, idx))

    if point_candidates:
        mapping["points_col"] = min(point_candidates)[1]
    return mapping


def _append_parsed_row(rows: list[dict], col_map: dict[str, int], cells: list[str]) -> None:
    if "rank_col" not in col_map or "name_col" not in col_map:
        return
    required_max = max(col_map["rank_col"], col_map["name_col"])
    if len(cells) <= required_max:
        return

    rank = _parse_rank(cells[col_map["rank_col"]])
    if rank is None:
        return

    name = _clean_text(cells[col_map["name_col"]])
    name_key = _header_key(name)
    if not name or (name_key and name_key in _SKIP_ROW_TOKENS):
        return

    club = None
    if "club_col" in col_map and col_map["club_col"] < len(cells):
        club = _clean_text(cells[col_map["club_col"]]) or None

    school = None
    if "school_col" in col_map and col_map["school_col"] < len(cells):
        school = _clean_text(cells[col_map["school_col"]]) or None

    points_idx = col_map.get("points_col")
    if points_idx is not None and points_idx < len(cells):
        points = _parse_points(cells[points_idx])
    elif len(cells) > max(col_map.values()) + 1:
        points = _parse_points(cells[-1])
    else:
        points = None

    metadata = {}
    if school:
        metadata["school"] = school

    row = {
        "rank": rank,
        "name": name,
        "club": club,
        "points": points,
    }
    if school:
        row["school"] = school
    if metadata:
        row["metadata"] = metadata
    rows.append(row)


def _parse_matrix(matrix: list[list[str]]) -> list[dict]:
    rows: list[dict] = []
    col_map: dict[str, int] | None = None
    seen: set[tuple[int, str]] = set()

    for raw_cells in matrix:
        cells = [_clean_text(cell) for cell in raw_cells]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue

        if col_map is None:
            candidate = _detect_columns(cells)
            if "rank_col" in candidate and "name_col" in candidate:
                col_map = candidate
                continue
            if len(cells) >= 2 and _parse_rank(cells[0]) is not None:
                col_map = {"rank_col": 0, "name_col": 1}
                if len(cells) >= 3:
                    col_map["club_col"] = 2
                if len(cells) >= 4:
                    col_map["points_col"] = len(cells) - 1
            else:
                continue

        before = len(rows)
        _append_parsed_row(rows, col_map, cells)
        if len(rows) == before:
            continue

        key = (rows[-1]["rank"], rows[-1]["name"])
        if key in seen:
            rows.pop()
        else:
            seen.add(key)

    return rows


def _html_tables_to_matrices(html: str) -> list[list[list[str]]]:
    soup = BeautifulSoup(html, "html.parser")
    matrices = []
    for table in soup.find_all("table"):
        matrix = []
        for tr in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in tr.find_all(["td", "th"])]
            if cells:
                matrix.append(cells)
        if matrix:
            matrices.append(matrix)
    return matrices


def _text_to_matrix(text: str) -> list[list[str]]:
    matrix = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" in line:
            cells = line.split("\t")
        elif "|" in line:
            cells = line.split("|")
        else:
            cells = re.split(r"\s{2,}", line)
        if len(cells) > 1:
            matrix.append(cells)
    return matrix


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Singapore federation ranking rows into rank/name/club/points dictionaries."""
    if not html_or_text or not html_or_text.strip():
        return []

    lowered = html_or_text.lower()
    if any(marker in lowered for marker in _NO_DATA_MARKERS):
        return []

    parsed_rows: list[dict] = []
    for matrix in _html_tables_to_matrices(html_or_text):
        parsed_rows.extend(_parse_matrix(matrix))
    if parsed_rows:
        return parsed_rows

    text = BeautifulSoup(html_or_text, "html.parser").get_text("\n", strip=True)
    return _parse_matrix(_text_to_matrix(text))


def _sheet_code(weapon: str, gender: str, category: str) -> str:
    category_code = {"senior": "S", "junior": "J"}.get(category.lower())
    gender_code = {"men": "M", "women": "W"}.get(gender.lower())
    weapon_code = {"foil": "F", "epee": "E", "sabre": "S"}.get(weapon.lower())
    if not category_code or not gender_code or not weapon_code:
        raise ValueError(f"unsupported Singapore ranking combo: {weapon} {gender} {category}")
    return f"{category_code}{gender_code}{weapon_code}"


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _extract_xlsx_sheet_text(content: bytes, sheet_code: str) -> str | None:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        print(f"    XLSX open error: {exc}")
        return None

    worksheet = None
    for candidate in workbook.worksheets:
        if candidate.title.strip().upper() == sheet_code.upper():
            worksheet = candidate
            break
    if worksheet is None:
        print(f"    Sheet {sheet_code} not found in Singapore ranking workbook")
        return None

    lines = []
    for row in worksheet.iter_rows(values_only=True):
        cells = [_format_cell(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        if cells and any(cell for cell in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _download_workbook(weapon: str) -> bytes | None:
    if weapon in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[weapon]

    page_url = DOWNLOAD_PAGES.get(weapon)
    if not page_url:
        print(f"    No Singapore download page configured for {weapon}")
        return None

    try:
        page = federation_request("get", page_url, headers=HEADERS, timeout=30, allow_redirects=True)
    except requests.RequestException as exc:
        print(f"    Request error for {page_url}: {exc}")
        return None
    if page.status_code != 200:
        print(f"    HTTP {page.status_code} for {page_url}")
        return None

    soup = BeautifulSoup(page.text, "html.parser")
    link = soup.select_one("[data-downloadurl]")
    if not link:
        print(f"    No data-downloadurl found on {page_url}")
        return None

    download_url = urljoin(page_url, link.get("data-downloadurl", ""))
    try:
        response = federation_request("get", download_url, headers=HEADERS, timeout=45, allow_redirects=True)
    except requests.RequestException as exc:
        print(f"    Request error for {download_url}: {exc}")
        return None
    if response.status_code != 200:
        print(f"    HTTP {response.status_code} for {download_url}")
        return None

    content_type = response.headers.get("content-type", "").lower()
    if "spreadsheetml.sheet" not in content_type and not response.content.startswith(b"PK\x03\x04"):
        print(f"    Unexpected Singapore ranking format from {download_url}: {content_type}")
        return None

    _WORKBOOK_CACHE[weapon] = response.content
    return response.content


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Fetch and extract ranking content for one Singapore weapon/gender/category combo."""
    try:
        sheet_code = _sheet_code(weapon, gender, category)
    except ValueError as exc:
        print(f"    {exc}")
        return None

    workbook = _download_workbook(weapon)
    if not workbook:
        return None
    return _extract_xlsx_sheet_text(workbook, sheet_code)


def current_season() -> str:
    """Return the current fencing season as YYYY-YYYY."""
    now = datetime.now(timezone.utc)
    start_year = now.year - 1 if now.month < 7 else now.year
    season = f"{start_year}-{start_year + 1}"
    try:
        import season_utils  # type: ignore

        if hasattr(season_utils, "normalize_season"):
            normalized = season_utils.normalize_season(season)
            if isinstance(normalized, str):
                return normalized
    except Exception:
        pass
    return season


def main() -> None:
    run_log = ScraperRunLogger("scrape_fed_sgp").start()
    season = current_season()
    previous_state = get_state(SOURCE, "last_run")
    if previous_state:
        print(f"Previous Singapore federation run state found: {previous_state}")

    print(f"Fencing Singapore rankings — season {season}")
    print(f"Ranking files page: {RANKING_FILES_PAGE}")

    total_written = 0
    total_failed = 0
    total_skipped = 0
    working_combos = 0
    failed_combos: list[str] = []

    try:
        for index, (weapon, gender, category) in enumerate(RANKING_COMBOS):
            combo_label = f"{weapon} {gender} {category}"
            print(f"  {combo_label}...")

            content = fetch_rankings_page(weapon, gender, category)
            if not content:
                total_failed += 1
                failed_combos.append(combo_label)
            else:
                parsed = parse_rankings_table(content)
                if not parsed:
                    print("    No rows parsed")
                    total_failed += 1
                    failed_combos.append(combo_label)
                else:
                    source_page = DOWNLOAD_PAGES.get(weapon)
                    rows = []
                    for row in parsed:
                        metadata = {
                            "source_page": source_page,
                            "source_files_page": RANKING_FILES_PAGE,
                            "source_format": "xlsx",
                            "sheet": _sheet_code(weapon, gender, category),
                        }
                        metadata.update(row.get("metadata") or {})
                        rows.append(
                            build_ranking_row(
                                source=SOURCE,
                                season=season,
                                weapon=weapon,
                                gender=gender,
                                category=category,
                                rank=row["rank"],
                                name=row["name"],
                                country=COUNTRY,
                                club=row.get("club"),
                                points=row.get("points"),
                                metadata=metadata,
                            )
                        )
                    written = write_rankings(rows, source=SOURCE, season=season)
                    print(f"    Parsed {len(rows)} rows; written {written} rows")
                    total_written += written
                    working_combos += 1

            if index < len(RANKING_COMBOS) - 1:
                time.sleep(REQUEST_DELAY)

        state = {
            "season": season,
            "written": total_written,
            "failed": total_failed,
            "skipped": total_skipped,
            "working_combos": working_combos,
            "total_combos": len(RANKING_COMBOS),
            "failed_combos": failed_combos,
            "source_files_page": RANKING_FILES_PAGE,
            "download_pages": DOWNLOAD_PAGES,
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }
        set_state(SOURCE, "last_run", state)
        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata=state,
        )
        if failed_combos:
            print(f"Failed combos: {', '.join(failed_combos)}")
        print(
            "Done — "
            f"written={total_written}, failed={total_failed}, skipped={total_skipped}, "
            f"working_combos={working_combos}/{len(RANKING_COMBOS)}"
        )
    except Exception as exc:
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
