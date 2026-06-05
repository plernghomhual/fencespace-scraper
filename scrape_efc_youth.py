"""
European Fencing Confederation youth circuit results scraper.

Probe summary (2026-06-02):
  * Current public EFC result index lives at https://www.fencing-efc.eu/results.
  * EFC detail pages and older mirrors expose public result rows with rank,
    points, name, age, and nationality columns, plus public document/result links.
  * Public downloads can be federation-hosted PDF/XLSX files or external result
    links. This scraper imports only public competition result rows and ignores
    minor profile links or age fields.
"""

from __future__ import annotations

import io
import os
import re
import time
import unicodedata
from datetime import date, datetime, timezone
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

from run_logger import ScraperRunLogger
from scraper_state import get_state, set_state

SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_SERVICE_KEY")

supabase = None
if SUPABASE_URL and SUPABASE_KEY:
    from supabase import create_client

    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

SOURCE = "efc_youth"
DEFAULT_SOURCE_URLS = ["https://www.fencing-efc.eu/results"]
REQUEST_DELAY = float(os.environ.get("EFC_YOUTH_REQUEST_DELAY", "1.5"))
BATCH_SIZE = 100

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; FenceSpace/1.0; +https://fencespace.app)",
    "Accept": "text/html,application/pdf,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*;q=0.8",
}

COUNTRY_ALIASES = {
    "ALBANIA": "ALB",
    "ANDORRA": "AND",
    "ARGENTINA": "ARG",
    "ARMENIA": "ARM",
    "AUSTRALIA": "AUS",
    "AUSTRIA": "AUT",
    "AZERBAIJAN": "AZE",
    "BELARUS": "BLR",
    "BELGIUM": "BEL",
    "BOSNIA AND HERZEGOVINA": "BIH",
    "BRAZIL": "BRA",
    "BULGARIA": "BUL",
    "CANADA": "CAN",
    "CHINA": "CHN",
    "CROATIA": "CRO",
    "CYPRUS": "CYP",
    "CZECH REPUBLIC": "CZE",
    "CZECHIA": "CZE",
    "DENMARK": "DEN",
    "EGYPT": "EGY",
    "ESTONIA": "EST",
    "FINLAND": "FIN",
    "FRANCE": "FRA",
    "GEORGIA": "GEO",
    "GERMANY": "GER",
    "GREAT BRITAIN": "GBR",
    "GREECE": "GRE",
    "HONG KONG": "HKG",
    "HUNGARY": "HUN",
    "ICELAND": "ISL",
    "IRELAND": "IRL",
    "ISRAEL": "ISR",
    "ITALY": "ITA",
    "JAPAN": "JPN",
    "KAZAKHSTAN": "KAZ",
    "KOREA": "KOR",
    "LATVIA": "LAT",
    "LITHUANIA": "LTU",
    "LUXEMBOURG": "LUX",
    "MALTA": "MLT",
    "MEXICO": "MEX",
    "MOLDOVA": "MDA",
    "MONTENEGRO": "MNE",
    "NETHERLANDS": "NED",
    "NORTH MACEDONIA": "MKD",
    "NORWAY": "NOR",
    "POLAND": "POL",
    "PORTUGAL": "POR",
    "ROMANIA": "ROU",
    "RUSSIA": "RUS",
    "SERBIA": "SRB",
    "SINGAPORE": "SGP",
    "SLOVAKIA": "SVK",
    "SLOVENIA": "SLO",
    "SPAIN": "ESP",
    "SWEDEN": "SWE",
    "SWITZERLAND": "SUI",
    "TURKEY": "TUR",
    "TURKIYE": "TUR",
    "UKRAINE": "UKR",
    "UNITED KINGDOM": "GBR",
    "UNITED STATES": "USA",
    "UNITED STATES OF AMERICA": "USA",
    "USA": "USA",
}

WEAPON_PATTERNS = [
    (r"\b(epee|degen|espada)\b", "Epee"),
    (r"\b(foil|fleuret|florett|fioretto|florete)\b", "Foil"),
    (r"\b(sabre|saber|sable|sabel|sciabola)\b", "Sabre"),
]
GENDER_PATTERNS = [
    (r"\b(women|woman|female|feminine|feminin|femmes|dames|girls|fille|filles|femenino|femminile)\b", "Women"),
    (r"\b(men|man|male|masculine|masculin|hommes|herr|herren|boys|garcon|garcons|masculino|maschile)\b", "Men"),
]

RANK_HEADERS = {"rank", "rang", "platz", "place", "pos", "position", "classement", "classification", "classifica", "clasificacion"}
POINT_HEADERS = {"points", "point", "pts", "punkte", "puntos", "punti"}
NAME_HEADERS = {"name", "nom", "nominativo", "fencer", "athlete", "competitor", "tireur", "tirador"}
COUNTRY_HEADERS = {"nationality", "nation", "country", "pays", "pais", "paese", "land"}
CLUB_HEADERS = {"club", "clubs", "verein", "societe", "association", "organization", "organisation"}
FIE_HEADERS = {"fie", "fie id", "id fie", "fencer id", "license", "licence"}
FIRST_HEADERS = {"first name", "given name", "prenom", "forename"}
FAMILY_HEADERS = {"family name", "last name", "surname", "nom de famille"}

UNMATCHED_FENCERS: list[dict] = []


class RateLimiter:
    def __init__(self, min_interval: float = REQUEST_DELAY):
        self.min_interval = max(0.0, float(min_interval))
        self._last_call = 0.0

    def wait(self):
        if self.min_interval <= 0:
            return
        now = time.monotonic()
        remaining = self.min_interval - (now - self._last_call)
        if remaining > 0:
            time.sleep(remaining)
        self._last_call = time.monotonic()


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_key(value) -> str:
    text = unicodedata.normalize("NFKD", clean_text(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^0-9A-Za-z]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def title_word(word: str) -> str:
    return "-".join(piece[:1].upper() + piece[1:].lower() for piece in word.split("-") if piece)


def title_name(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    return " ".join(title_word(part) for part in text.split())


def normalize_person_name(value) -> str | None:
    text = clean_text(value).strip(",")
    if not text:
        return None
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        return title_name(f"{first} {last}")

    parts = text.split()
    upper_prefix = 0
    while upper_prefix < len(parts):
        letters = re.sub(r"[^A-Za-z]", "", parts[upper_prefix])
        if letters and letters.upper() == letters:
            upper_prefix += 1
            continue
        break
    if 0 < upper_prefix < len(parts):
        return title_name(f"{' '.join(parts[upper_prefix:])} {' '.join(parts[:upper_prefix])}")
    return title_name(text)


def normalize_country(value) -> str | None:
    text = clean_text(value)
    if not text:
        return None
    key = normalize_key(text).upper()
    if re.fullmatch(r"[A-Z]{3}", key):
        return key
    return COUNTRY_ALIASES.get(key) or title_name(text)


def parse_rank(value) -> int | None:
    match = re.search(r"\d+", clean_text(value))
    return int(match.group(0)) if match else None


def parse_points(value) -> float | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", ".")
    text = re.sub(r"[^0-9.+-]", "", text)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_fie_id(value) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = clean_text(value)
    if not text:
        return None
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text or None


def parse_date_range(value) -> tuple[str | None, str | None]:
    if isinstance(value, (datetime, date)):
        iso = value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
        return iso, iso

    text = clean_text(value)
    if not text:
        return None, None

    matches = re.findall(r"\b(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})\b", text)
    dates = []
    for day, month, year in matches[:2]:
        year = f"20{year}" if len(year) == 2 else year
        try:
            dates.append(datetime(int(year), int(month), int(day)).date().isoformat())
        except ValueError:
            continue
    if dates:
        return dates[0], dates[-1]

    iso_matches = re.findall(r"\b((?:19|20)\d{2})-(\d{1,2})-(\d{1,2})\b", text)
    for year, month, day in iso_matches[:2]:
        try:
            dates.append(datetime(int(year), int(month), int(day)).date().isoformat())
        except ValueError:
            continue
    if dates:
        return dates[0], dates[-1]
    return None, None


def classify_event(label: str, page_category: str | None = None) -> dict:
    normalized = normalize_key(label)

    category = None
    for category_text in [normalized, normalize_key(page_category or "")]:
        under_age = re.search(r"\bu\s*([0-9]{2})\b", category_text)
        if under_age:
            age = int(under_age.group(1))
            if age <= 15:
                category = f"U{age}"
            elif age <= 17:
                category = "Cadet"
            elif age <= 20:
                category = "Junior"
            else:
                category = f"U{age}"
        elif re.search(r"\b(junior|juniors)\b", category_text):
            category = "Junior"
        elif re.search(r"\b(cadet|cadets)\b", category_text):
            category = "Cadet"
        elif re.search(r"\b(minime|minimes|benjamin|benjamins)\b", category_text):
            category = "U14"
        if category:
            break

    weapon = None
    for pattern, value in WEAPON_PATTERNS:
        if re.search(pattern, normalized):
            weapon = value
            break

    gender = None
    for pattern, value in GENDER_PATTERNS:
        if re.search(pattern, normalized):
            gender = value
            break

    team = bool(re.search(r"\b(team|teams|equipe|equipes|mannschaft)\b", normalized))
    return {"category": category, "weapon": weapon, "gender": gender, "team": team}


def allowed_category(category: str | None) -> bool:
    return category in {"Cadet", "Junior"}


def slugify(value: str) -> str:
    key = normalize_key(value)
    return re.sub(r"[^a-z0-9]+", "-", key).strip("-") or "event"


def source_id_for_event(source_url: str, event_name: str) -> str:
    parsed = urlparse(source_url)
    path = parsed.path.strip("/").replace("/", ":") or parsed.netloc
    return f"{SOURCE}:{slugify(path)}:{slugify(event_name)}"


def extract_page_metadata(soup: BeautifulSoup) -> dict:
    metadata = {}
    for dt in soup.find_all("dt"):
        dd = dt.find_next_sibling("dd")
        if not dd:
            continue
        key = normalize_key(dt.get_text(" ", strip=True)).rstrip(":")
        metadata[key] = clean_text(dd.get_text(" ", strip=True))

    text = soup.get_text("\n", strip=True)
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    for index, line in enumerate(lines):
        key_match = re.match(r"^(date|place|category|weapon|event|competition)\s*:?\s*(.*)$", normalize_key(line))
        if not key_match:
            continue
        key = key_match.group(1)
        value = clean_text(line.split(":", 1)[1]) if ":" in line else None
        if not value and index + 1 < len(lines):
            value = lines[index + 1]
        if value and key not in metadata:
            metadata[key] = value
    return metadata


def extract_source_links(soup: BeautifulSoup, base_url: str) -> list[str]:
    links = []
    seen = set()
    for link in soup.find_all("a", href=True):
        href = urljoin(base_url, link["href"])
        label = f"{link.get_text(' ', strip=True)} {href}"
        if not re.search(r"pdf|xlsx?|download|document|result|live|engarde", label, re.I):
            continue
        if href not in seen:
            seen.add(href)
            links.append(href)
    return links


def _header_name(raw_header: str) -> str | None:
    key = normalize_key(raw_header)
    if key in RANK_HEADERS:
        return "rank"
    if key in POINT_HEADERS:
        return "points"
    if key in NAME_HEADERS:
        return "name"
    if key in COUNTRY_HEADERS:
        return "country"
    if key in CLUB_HEADERS:
        return "club"
    if key in FIE_HEADERS or ("fie" in key and "id" in key):
        return "fie_id"
    if key in FIRST_HEADERS or ("first" in key and "name" in key):
        return "first_name"
    if key in FAMILY_HEADERS or "family" in key or "surname" in key or ("last" in key and "name" in key):
        return "family_name"
    return None


def column_map(headers: list[str]) -> dict:
    mapping = {}
    for index, header in enumerate(headers):
        name = _header_name(header)
        if name and name not in mapping:
            mapping[name] = index
    return mapping


def _get_cell(cells: list[str], mapping: dict, field: str):
    index = mapping.get(field)
    if index is None or index >= len(cells):
        return None
    return cells[index]


def _looks_like_header(cells: list[str]) -> bool:
    mapping = column_map(cells)
    has_name = "name" in mapping or ("first_name" in mapping and "family_name" in mapping)
    return "rank" in mapping and has_name


def _row_name(cells: list[str], mapping: dict) -> str | None:
    raw_name = _get_cell(cells, mapping, "name")
    if raw_name:
        return normalize_person_name(raw_name)
    first = _get_cell(cells, mapping, "first_name")
    family = _get_cell(cells, mapping, "family_name")
    if first or family:
        return normalize_person_name(f"{first or ''} {family or ''}")
    return None


def parse_result_matrix(matrix: list[list], event: dict, source_url: str) -> list[dict]:
    rows: list[dict] = []
    header_index = None
    mapping = {}
    for index, raw_row in enumerate(matrix):
        cells = [clean_text(cell) for cell in raw_row]
        if _looks_like_header(cells):
            header_index = index
            mapping = column_map(cells)
            break
    if header_index is None:
        return rows

    for raw_row in matrix[header_index + 1 :]:
        cells = [clean_text(cell) for cell in raw_row]
        if not any(cells):
            continue
        if _looks_like_header(cells):
            continue
        rank = parse_rank(_get_cell(cells, mapping, "rank"))
        name = _row_name(cells, mapping)
        if rank is None or not name:
            continue
        rows.append(
            {
                "rank": rank,
                "fencer": name,
                "country": normalize_country(_get_cell(cells, mapping, "country")),
                "club": clean_text(_get_cell(cells, mapping, "club")) or None,
                "points": parse_points(_get_cell(cells, mapping, "points")),
                "fie_id": normalize_fie_id(_get_cell(cells, mapping, "fie_id")),
                "source_url": source_url,
                "date": event.get("date"),
            }
        )
    return rows


def _previous_heading(table) -> str | None:
    heading = table.find_previous(["h2", "h3", "h4", "h5", "h6"])
    if not heading:
        heading = table.find_previous(["h1"])
    return clean_text(heading.get_text(" ", strip=True)) if heading else None


def _event_from_table(table, *, source_url: str, metadata: dict, competition_name: str, source_links: list[str]) -> tuple[dict | None, dict | None]:
    event_name = _previous_heading(table) or metadata.get("event") or competition_name
    classification = classify_event(event_name, metadata.get("category"))
    if not allowed_category(classification["category"]):
        return None, {
            "source_url": source_url,
            "event_name": event_name,
            "reason": f"blocked_minor_category:{classification['category'] or 'unknown'}",
        }
    if not classification["weapon"] or not classification["gender"]:
        return None, {
            "source_url": source_url,
            "event_name": event_name,
            "reason": "unclassifiable_event",
        }

    start_date, end_date = parse_date_range(metadata.get("date"))
    matrix = []
    for tr in table.find_all("tr"):
        matrix.append([cell.get_text(" ", strip=True) for cell in tr.find_all(["th", "td"])])

    event = {
        "source_id": source_id_for_event(source_url, event_name),
        "competition_name": competition_name,
        "event_name": event_name,
        "category": classification["category"],
        "weapon": classification["weapon"],
        "gender": classification["gender"],
        "team": classification["team"],
        "date": start_date,
        "end_date": end_date,
        "place": metadata.get("place"),
        "source_url": source_url,
        "source_links": source_links,
    }
    event["results"] = parse_result_matrix(matrix, event, source_url)
    return event if event["results"] else None, None


def parse_event_page(html: str, source_url: str) -> dict:
    soup = BeautifulSoup(html or "", "html.parser")
    metadata = extract_page_metadata(soup)
    h1 = soup.find("h1")
    competition_name = clean_text(h1.get_text(" ", strip=True)) if h1 else metadata.get("competition") or "EFC Youth Circuit"
    source_links = extract_source_links(soup, source_url)

    events = []
    skipped = []
    for table in soup.find_all("table"):
        event, skip = _event_from_table(
            table,
            source_url=source_url,
            metadata=metadata,
            competition_name=competition_name,
            source_links=source_links,
        )
        if event:
            events.append(event)
        elif skip and skip not in skipped:
            skipped.append(skip)
    return {"events": events, "skipped": skipped}


def _country_alias_patterns() -> list[tuple[str, str]]:
    aliases = list(COUNTRY_ALIASES.items())
    for code in sorted(set(COUNTRY_ALIASES.values())):
        aliases.append((code, code))
    aliases.sort(key=lambda item: len(item[0]), reverse=True)
    return aliases


def _split_country_tail(text: str) -> tuple[str, str | None]:
    value = clean_text(text)
    for alias, code in _country_alias_patterns():
        pattern = rf"^(?P<name>.+?)(?:\s+\d{{1,2}})?\s+{re.escape(alias)}$"
        match = re.match(pattern, value, re.I)
        if match:
            return clean_text(match.group("name")), code
    match = re.match(r"^(?P<name>.+?)(?:\s+\d{1,2})?\s+(?P<country>[A-Z]{3})$", value)
    if match:
        return clean_text(match.group("name")), normalize_country(match.group("country"))
    return value, None


def _parse_text_result_line(line: str, event: dict, source_url: str) -> dict | None:
    match = re.match(r"^\s*(?P<rank>\d{1,4}(?:=|\.)?)\s+(?P<points>-?\d+(?:[,.]\d+)?)\s+(?P<rest>.+?)\s*$", line)
    if not match:
        return None
    name_part, country = _split_country_tail(match.group("rest"))
    name = normalize_person_name(name_part)
    if not name:
        return None
    return {
        "rank": parse_rank(match.group("rank")),
        "fencer": name,
        "country": country,
        "club": None,
        "points": parse_points(match.group("points")),
        "fie_id": None,
        "source_url": source_url,
        "date": event.get("date"),
    }


def parse_pdf_text_events(text: str, source_url: str) -> list[dict]:
    lines = [clean_text(line) for line in (text or "").splitlines() if clean_text(line)]
    if not lines:
        return []

    competition_name = lines[0]
    metadata = {"competition": competition_name}
    for index, line in enumerate(lines):
        key = normalize_key(line).rstrip(":")
        if key in {"date", "place"} and index + 1 < len(lines):
            metadata[key] = lines[index + 1]
        elif ":" in line:
            key_part, value = line.split(":", 1)
            normalized_key = normalize_key(key_part)
            if normalized_key in {"date", "place", "competition"}:
                metadata[normalized_key] = clean_text(value)

    start_date, end_date = parse_date_range(metadata.get("date"))
    events = []
    current = None
    reading_rows = False

    for line in lines:
        classification = classify_event(line, metadata.get("category"))
        if allowed_category(classification["category"]) and classification["weapon"] and classification["gender"]:
            current = {
                "source_id": source_id_for_event(source_url, line),
                "competition_name": competition_name,
                "event_name": line,
                "category": classification["category"],
                "weapon": classification["weapon"],
                "gender": classification["gender"],
                "team": classification["team"],
                "date": start_date,
                "end_date": end_date,
                "place": metadata.get("place"),
                "source_url": source_url,
                "source_links": [source_url],
                "results": [],
            }
            events.append(current)
            reading_rows = False
            continue
        if current and _looks_like_header(line.split()):
            reading_rows = True
            continue
        if current and reading_rows:
            row = _parse_text_result_line(line, current, source_url)
            if row:
                current["results"].append(row)

    return [event for event in events if event["results"]]


def parse_xlsx_events(content: bytes, source_url: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    events = []
    for sheet in workbook.worksheets:
        matrix = [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        metadata = {}
        for row in matrix[:12]:
            if len(row) < 2:
                continue
            key = normalize_key(row[0])
            if key in {"competition", "date", "event", "place", "category"} and row[1]:
                metadata[key] = row[1]
        event_name = clean_text(metadata.get("event")) or sheet.title
        competition_name = clean_text(metadata.get("competition")) or "EFC Youth Circuit"
        classification = classify_event(event_name, metadata.get("category"))
        if not allowed_category(classification["category"]) or not classification["weapon"] or not classification["gender"]:
            continue
        start_date, end_date = parse_date_range(metadata.get("date"))
        event = {
            "source_id": source_id_for_event(source_url, f"{sheet.title}:{event_name}"),
            "competition_name": competition_name,
            "event_name": event_name,
            "category": classification["category"],
            "weapon": classification["weapon"],
            "gender": classification["gender"],
            "team": classification["team"],
            "date": start_date,
            "end_date": end_date,
            "place": clean_text(metadata.get("place")) or None,
            "source_url": source_url,
            "source_links": [source_url],
        }
        event["results"] = parse_result_matrix(matrix, event, source_url)
        if event["results"]:
            events.append(event)
    return events


def parse_results_index(html: str, base_url: str) -> list[str]:
    soup = BeautifulSoup(html or "", "html.parser")
    urls = []
    seen = set()
    for link in soup.find_all("a", href=True):
        href = urljoin(base_url, link["href"])
        label = f"{link.get_text(' ', strip=True)} {href}"
        if "/results" not in urlparse(href).path:
            continue
        classification = classify_event(label)
        if not allowed_category(classification["category"]):
            continue
        if href not in seen:
            seen.add(href)
            urls.append(href)
    return urls


def blocked_source_stub(url: str, reason: str) -> dict:
    return {"source_url": url, "reason": reason}


def fetch_source_content(url: str):
    try:
        response = requests.get(url, headers=HEADERS, timeout=30)
    except Exception as exc:
        return None, blocked_source_stub(url, str(exc))
    if response.status_code in {401, 403}:
        return None, blocked_source_stub(url, f"HTTP {response.status_code}")
    if response.status_code == 404:
        return None, blocked_source_stub(url, "HTTP 404")
    if response.status_code != 200:
        return None, blocked_source_stub(url, f"HTTP {response.status_code}")
    return {
        "url": url,
        "content": response.content,
        "text": response.text,
        "content_type": response.headers.get("content-type", ""),
    }, None


def _extract_pdf_text(content: bytes) -> str:
    import pdfplumber

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        return "\n".join(page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages)


def parse_source_content(source: dict) -> dict:
    url = source["url"]
    content_type = source.get("content_type", "").lower()
    lower_url = url.lower()
    content = source.get("content") or b""
    if lower_url.endswith((".xlsx", ".xlsm")) or "spreadsheet" in content_type:
        return {"events": parse_xlsx_events(content, url), "skipped": []}
    if lower_url.endswith(".pdf") or "pdf" in content_type or content.startswith(b"%PDF"):
        return {"events": parse_pdf_text_events(_extract_pdf_text(content), url), "skipped": []}
    return parse_event_page(source.get("text") or content.decode("utf-8", errors="ignore"), url)


def upsert_tournament(event: dict):
    if supabase is None:
        return None
    row = {
        "source_id": event["source_id"],
        "name": f"{event['competition_name']} - {event['event_name']}",
        "season": (event.get("date") or "")[:4] or None,
        "type": "efc_youth_circuit",
        "weapon": event["weapon"],
        "gender": event["gender"],
        "category": event["category"],
        "country": None,
        "start_date": event.get("date"),
        "end_date": event.get("end_date"),
        "has_results": True,
        "metadata": {
            "source": SOURCE,
            "source_url": event.get("source_url"),
            "source_links": event.get("source_links", []),
            "competition_name": event.get("competition_name"),
            "event_name": event.get("event_name"),
            "team": event.get("team", False),
        },
    }
    try:
        result = supabase.table("fs_tournaments").upsert(row, on_conflict="source_id").execute()
        return result.data[0]["id"] if result.data else None
    except Exception as exc:
        print(f"  Tournament upsert failed for {event['source_id']}: {exc}")
        return None


def match_fencer(result_row: dict):
    if supabase is None:
        return None
    fie_id = result_row.get("fie_id")
    if fie_id:
        try:
            rows = (
                supabase.table("fs_fencers")
                .select("id")
                .eq("fie_id", str(fie_id))
                .limit(2)
                .execute()
                .data
                or []
            )
            if rows:
                return rows[0]["id"]
        except Exception:
            pass
    if result_row.get("fencer") and result_row.get("country"):
        try:
            rows = (
                supabase.table("fs_fencers")
                .select("id")
                .ilike("name", result_row["fencer"])
                .eq("country", result_row["country"])
                .limit(2)
                .execute()
                .data
                or []
            )
            return rows[0]["id"] if len(rows) == 1 else None
        except Exception:
            return None
    return None


def log_unmatched_fencer(item: dict):
    UNMATCHED_FENCERS.append(item)
    print(f"  Unmatched EFC youth fencer: {item.get('name')} {item.get('country') or ''}".strip())


def build_result_rows(tournament_id, event: dict, matcher=None, unmatched_logger=None) -> list[dict]:
    matcher = matcher or match_fencer
    unmatched_logger = unmatched_logger or log_unmatched_fencer
    rows = []
    for result in event.get("results", []):
        rank = parse_rank(result.get("rank"))
        name = result.get("fencer")
        if rank is None or not name:
            continue
        fencer_id = matcher(result)
        unmatched = {"name": name, "country": result.get("country"), "fie_id": result.get("fie_id")}
        metadata = {
            "source": SOURCE,
            "source_url": event.get("source_url"),
            "event_name": event.get("event_name"),
            "category": event.get("category"),
            "weapon": event.get("weapon"),
            "gender": event.get("gender"),
            "club": result.get("club"),
            "points": result.get("points"),
            "fie_id": result.get("fie_id"),
        }
        if not fencer_id:
            metadata["unmatched"] = True
            unmatched_logger(unmatched)
        row = {
            "tournament_id": tournament_id,
            "name": name,
            "nationality": result.get("country"),
            "rank": rank,
            "fencer_id": fencer_id,
            "metadata": metadata,
        }
        if result.get("fie_id") and str(result["fie_id"]).isdigit():
            row["fie_fencer_id"] = int(result["fie_id"])
        rows.append(row)
    return rows


def upsert_results(tournament_id, event: dict) -> int:
    if supabase is None:
        return 0
    rows = build_result_rows(tournament_id, event)
    if not rows:
        return 0
    supabase.table("fs_results").delete().eq("tournament_id", tournament_id).execute()
    written = 0
    for index in range(0, len(rows), BATCH_SIZE):
        batch = rows[index : index + BATCH_SIZE]
        try:
            supabase.table("fs_results").insert(batch).execute()
            written += len(batch)
        except Exception as exc:
            print(f"  Results insert failed for {tournament_id}: {exc}")
    return written if written == len(rows) else 0


def _source_urls_from_env() -> list[str]:
    configured = [url.strip() for url in os.environ.get("EFC_YOUTH_SOURCES", "").split(",") if url.strip()]
    return configured or list(DEFAULT_SOURCE_URLS)


def run_once(source_urls: list[str] | None = None, run_logger=None) -> dict:
    logger = (run_logger or ScraperRunLogger("scrape_efc_youth")).start()
    UNMATCHED_FENCERS.clear()
    blocked_sources = []
    written = failed = skipped = 0
    limiter = RateLimiter()
    done_source_ids = set(get_state(SOURCE, "done_source_ids") or [])
    queue = list(source_urls or _source_urls_from_env())
    seen_urls = set()

    try:
        while queue:
            url = queue.pop(0)
            if url in seen_urls:
                continue
            seen_urls.add(url)
            limiter.wait()
            source, blocked = fetch_source_content(url)
            if blocked:
                blocked_sources.append(blocked)
                skipped += 1
                continue

            parsed = parse_source_content(source)
            events = parsed.get("events", [])
            if not events:
                discovered = parse_results_index(source.get("text") or "", source["url"])
                queue.extend(link for link in discovered if link not in seen_urls)
                skipped += len(parsed.get("skipped", []))
                if not discovered and not parsed.get("skipped"):
                    skipped += 1
                continue

            skipped += len(parsed.get("skipped", []))
            for event in events:
                source_id = event["source_id"]
                if source_id in done_source_ids:
                    skipped += 1
                    continue
                tournament_id = upsert_tournament(event)
                if not tournament_id:
                    failed += 1
                    continue
                count = upsert_results(tournament_id, event)
                if count == 0:
                    failed += 1
                    continue
                written += count
                done_source_ids.add(source_id)
                set_state(SOURCE, "done_source_ids", sorted(done_source_ids))

        result = {"written": written, "failed": failed, "skipped": skipped}
        set_state(SOURCE, "last_unmatched_fencers", list(UNMATCHED_FENCERS))
        set_state(SOURCE, "last_run_summary", result)
        logger.complete(
            written=written,
            failed=failed,
            skipped=skipped,
            metadata={"blocked_sources": blocked_sources, "unmatched_fencers": list(UNMATCHED_FENCERS)},
        )
        return result
    except Exception as exc:
        logger.error(str(exc))
        raise


def main():
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("SUPABASE_URL and SUPABASE_SERVICE_KEY must be set.")
    print(f"EFC youth scraper starting - {datetime.now(timezone.utc).isoformat()}")
    result = run_once()
    print(f"Done - written={result['written']}, failed={result['failed']}, skipped={result['skipped']}")


if __name__ == "__main__":
    main()
