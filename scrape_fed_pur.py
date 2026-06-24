"""
scrape_fed_pur.py - Puerto Rico national federation rankings scraper.

Probe findings, 2026-06-02:
  - Requested host fepur.org did not resolve from the local sandbox probe.
  - Current public federation site: https://fedesgrimapuertorico.org/ranking/
  - Request method: GET.
  - Ranking page response format: public HTML with Spanish category sections.
  - Ranking files response format: XLSX workbook links from the ranking page.
  - Observed Adulto workbook link:
      /wp-content/uploads/2026/04/
      Ranking-Nacional-Adulto-2025-2026-Actualizado-abril-252026.xlsx
  - Browser probe exposed Adulto/Cadete links. Junior coverage is attempted
    from the public ranking page and logged as failed if no durable link exists.
"""

from __future__ import annotations

import io
import re
import time
import unicodedata
from datetime import UTC, datetime, timezone
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import set_state

SOURCE = "pur_fencing"
COUNTRY = "Puerto Rico"
COUNTRY_CODE = "pur"
BASE_URL = "https://fedesgrimapuertorico.org"
RANKING_PAGE = f"{BASE_URL}/ranking/"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    ),
    "Accept-Language": "es-PR,es;q=0.9,en;q=0.8",
    "Referer": BASE_URL,
}

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

_KNOWN_CATEGORY_URLS = {
    "Senior": (
        f"{BASE_URL}/wp-content/uploads/2026/04/"
        "Ranking-Nacional-Adulto-2025-2026-Actualizado-abril-252026.xlsx"
    ),
}
_CATEGORY_LINK_CACHE: dict[str, str] | None = None
_WORKBOOK_CACHE: dict[str, bytes] = {}

_RANK_HEADERS = {
    "#",
    "no",
    "num",
    "numero",
    "pos",
    "posicion",
    "puesto",
    "rank",
    "ranking",
    "lugar",
    "clasificacion",
}
_NAME_HEADERS = {"nombre", "atleta", "tirador", "tiradora", "deportista", "fencer", "name"}
_CLUB_HEADERS = {"club", "sala", "equipo", "escuela", "academia", "representacion"}
_POINT_HEADERS = {
    "puntos",
    "puntuacion",
    "puntaje",
    "pts",
    "points",
    "total",
    "totalpuntos",
    "puntosacumulados",
    "rankingpoints",
    "finalrankingpoints",
}
_SKIP_TOKENS = {
    "",
    "dns",
    "dnf",
    "dq",
    "dsq",
    "wd",
    "ret",
    "baja",
    "total",
    "totales",
    "resumen",
    "summary",
    "subtotal",
    "na",
    "n/a",
}
_NO_DATA_MARKERS = {
    "no hay ranking disponible",
    "no rankings available",
    "no ranking available",
    "no data",
    "sin ranking",
    "ranking no disponible",
}
_UNAVAILABLE_MARKERS = {
    "iniciar sesion",
    "iniciar sesión",
    "wp-login",
    "login",
    "captcha",
    "access denied",
    "forbidden",
    "cloudflare",
    "loading rankings",
}
_SPREADSHEET_EXTENSIONS = (".xlsx", ".xls", ".xlsm")

_CATEGORY_ALIASES = {
    "Senior": {"adulto", "adulta", "adultos", "senior", "mayor", "mayores"},
    "Junior": {"juvenil", "juveniles", "junior", "juniors", "u20", "sub20"},
}
_WEAPON_ALIASES = {
    "Foil": {"florete", "foil", "floret"},
    "Epee": {"espada", "epee", "épée"},
    "Sabre": {"sable", "sabre", "saber"},
}
_GENDER_ALIASES = {
    "Men": {"masculino", "masc", "hombre", "hombres", "varon", "varones", "varonil", "male", "men"},
    "Women": {"femenino", "fem", "mujer", "mujeres", "damas", "female", "women"},
}


def current_season() -> str:
    """Return the current fencing season as YYYY-YYYY, using season_utils if present."""
    now = datetime.now(UTC)
    season_end_year = now.year if now.month < 7 else now.year + 1
    try:
        import season_utils  # type: ignore

        if hasattr(season_utils, "normalize_season"):
            return season_utils.normalize_season(season_end_year)
        if hasattr(season_utils, "season_to_string"):
            return season_utils.season_to_string(season_end_year)
    except Exception:
        pass
    return f"{season_end_year - 1:04d}-{season_end_year:04d}"


def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _strip_accents(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in decomposed if not unicodedata.combining(ch))


def _token(value: str) -> str:
    value = _strip_accents(_clean_text(value)).lower()
    value = value.replace("#", " # ")
    value = value.replace("&", " and ").replace("/", " ")
    return re.sub(r"[^a-z0-9#]+", "", value)


_SKIP_NORMALIZED = {_token(item) for item in _SKIP_TOKENS}


def _normalised_words(value: str) -> set[str]:
    value = _strip_accents(_clean_text(value)).lower().replace("/", " ")
    return {word for word in re.split(r"[^a-z0-9]+", value) if word}


def _is_skip_value(value: str) -> bool:
    cleaned = _clean_text(value).lower()
    return cleaned in _SKIP_TOKENS or _token(cleaned) in _SKIP_NORMALIZED


def _parse_rank(value: str) -> int | None:
    text = _clean_text(value)
    if not text or _is_skip_value(text):
        return None
    match = re.match(r"^\s*(\d+)", text.rstrip("."))
    if not match:
        return None
    rank = int(match.group(1))
    return rank if rank > 0 else None


def _parse_points(value: str) -> float | None:
    text = _clean_text(value)
    if not text or _is_skip_value(text) or text in {"-", "—", "–"}:
        return None
    text = re.sub(r"[^0-9,.\-]", "", text)
    if text in {"", "-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        head, tail = text.rsplit(",", 1)
        if len(tail) in (1, 2):
            text = f"{head.replace(',', '')}.{tail}"
        else:
            text = text.replace(",", "")
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 2 and all(len(part) == 3 for part in parts[1:]):
            text = "".join(parts)
        elif len(parts) == 2 and len(parts[1]) == 3 and len(parts[0].lstrip("-")) <= 3:
            text = "".join(parts)

    try:
        return float(text)
    except ValueError:
        return None


def _detect_columns(cells: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cells):
        key = _token(cell)
        if key in _RANK_HEADERS and "rank" not in mapping:
            mapping["rank"] = idx
        elif key in _NAME_HEADERS and "name" not in mapping:
            mapping["name"] = idx
        elif key in _CLUB_HEADERS and "club" not in mapping:
            mapping["club"] = idx
        elif key in _POINT_HEADERS and "points" not in mapping:
            mapping["points"] = idx
        elif key.endswith("puntos") and "points" not in mapping:
            mapping["points"] = idx
        elif key.endswith("points") and "points" not in mapping:
            mapping["points"] = idx
    return mapping


def _append_parsed_row(rows: list[dict], mapping: dict[str, int], cells: list[str]) -> None:
    if not {"rank", "name"}.issubset(mapping):
        return
    if len(cells) <= max(mapping["rank"], mapping["name"]):
        return

    rank = _parse_rank(cells[mapping["rank"]])
    if rank is None:
        return

    name = _clean_text(cells[mapping["name"]])
    if not name or _is_skip_value(name):
        return

    club = None
    if "club" in mapping and mapping["club"] < len(cells):
        club = _clean_text(cells[mapping["club"]]) or None

    points = None
    if "points" in mapping and mapping["points"] < len(cells):
        points = _parse_points(cells[mapping["points"]])
    elif len(cells) >= 4:
        points = _parse_points(cells[-1])

    rows.append({"rank": rank, "name": name, "club": club, "points": points})


def _parse_matrix(matrix: list[list[str]]) -> list[dict]:
    parsed: list[dict] = []
    mapping: dict[str, int] | None = None
    seen: set[tuple[int, str]] = set()

    for raw_cells in matrix:
        cells = [_clean_text(cell) for cell in raw_cells]
        while cells and not cells[-1]:
            cells.pop()
        if not cells:
            continue

        candidate = _detect_columns(cells)
        if {"rank", "name"}.issubset(candidate):
            mapping = candidate
            continue

        if mapping is None:
            if len(cells) >= 2 and _parse_rank(cells[0]) is not None:
                mapping = {"rank": 0, "name": 1}
                if len(cells) >= 3:
                    mapping["club"] = 2
                if len(cells) >= 4:
                    mapping["points"] = len(cells) - 1
            else:
                continue

        before = len(parsed)
        _append_parsed_row(parsed, mapping, cells)
        if len(parsed) == before:
            continue
        key = (parsed[-1]["rank"], parsed[-1]["name"])
        if key in seen:
            parsed.pop()
        else:
            seen.add(key)

    return parsed


def _html_tables_to_matrices(html: str) -> list[list[list[str]]]:
    soup = BeautifulSoup(html, "html.parser")
    matrices = []
    for table in soup.find_all("table"):
        matrix = []
        for row in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["td", "th"])]
            if cells:
                matrix.append(cells)
        if matrix:
            matrices.append(matrix)
    return matrices


def _text_to_matrix(text: str) -> list[list[str]]:
    matrix = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "\t" in line:
            cells = line.split("\t")
        elif "|" in line:
            cells = line.split("|")
        else:
            cells = re.split(r"\s{2,}", line)
        if len(cells) > 1:
            matrix.append(cells)
    return matrix


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Puerto Rico ranking rows into rank/name/club/points dictionaries."""
    if not html_or_text or not html_or_text.strip():
        return []

    lowered = _strip_accents(html_or_text).lower()
    if any(marker in lowered for marker in _NO_DATA_MARKERS):
        return []

    parsed_rows: list[dict] = []
    for matrix in _html_tables_to_matrices(html_or_text):
        parsed_rows.extend(_parse_matrix(matrix))
    if parsed_rows:
        return parsed_rows

    text = BeautifulSoup(html_or_text, "html.parser").get_text("\n", strip=True)
    return _parse_matrix(_text_to_matrix(text))


def _category_from_text(value: str) -> str | None:
    words = _normalised_words(value)
    for category, aliases in _CATEGORY_ALIASES.items():
        if words.intersection(aliases):
            return category
    return None


def _is_spreadsheet_url(url: str) -> bool:
    return url.lower().split("?", 1)[0].endswith(_SPREADSHEET_EXTENSIONS)


def _extract_category_links(html: str, *, base_url: str = RANKING_PAGE) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    links: dict[str, str] = {}

    for anchor in soup.find_all("a", href=True):
        text = _clean_text(anchor.get_text(" ", strip=True))
        href = urljoin(base_url, anchor["href"])
        previous_heading = anchor.find_previous(["h1", "h2", "h3", "h4", "h5", "h6", "strong"])
        context = " ".join(
            part
            for part in [
                _clean_text(previous_heading.get_text(" ", strip=True)) if previous_heading else "",
                text,
                href,
            ]
            if part
        )
        category = _category_from_text(context)
        if category not in {"Senior", "Junior"}:
            continue
        if not (
            _is_spreadsheet_url(href)
            or "ranking" in _strip_accents(text).lower()
            or "ranking" in href.lower()
            or "download" in href.lower()
            or "ver ranking" in _strip_accents(text).lower()
        ):
            continue
        links.setdefault(category, href)

    return links


def _looks_unavailable(text: str) -> bool:
    normalised = _strip_accents(text).lower()
    return any(marker in normalised for marker in _UNAVAILABLE_MARKERS)


def _discover_category_links() -> dict[str, str]:
    global _CATEGORY_LINK_CACHE
    if _CATEGORY_LINK_CACHE is not None:
        return _CATEGORY_LINK_CACHE

    links = dict(_KNOWN_CATEGORY_URLS)
    try:
        response = federation_request(
            "get",
            RANKING_PAGE,
            headers=HEADERS,
            timeout=25,
            allow_redirects=True,
        )
    except requests.RequestException as exc:
        print(f"    Ranking page request error for {RANKING_PAGE}: {exc}")
        _CATEGORY_LINK_CACHE = links
        return links

    if response.status_code != 200:
        print(f"    No scrapeable rankings at {RANKING_PAGE} (HTTP {response.status_code})")
        _CATEGORY_LINK_CACHE = links
        return links

    if _looks_unavailable(response.text) and not _extract_category_links(response.text, base_url=response.url):
        print(f"    No scrapeable rankings at {response.url}: blocked/login/JS-only page")
        _CATEGORY_LINK_CACHE = links
        return links

    links.update(_extract_category_links(response.text, base_url=response.url))
    _CATEGORY_LINK_CACHE = links
    return links


def _spreadsheet_response(response, url: str) -> bool:
    content_type = response.headers.get("content-type", "").lower()
    return (
        _is_spreadsheet_url(url)
        or "spreadsheet" in content_type
        or "excel" in content_type
        or response.content.startswith(b"PK")
    )


def _first_spreadsheet_link(html: str, *, base_url: str) -> str | None:
    soup = BeautifulSoup(html, "html.parser")
    for anchor in soup.find_all("a", href=True):
        href = urljoin(base_url, anchor["href"])
        text = _clean_text(anchor.get_text(" ", strip=True))
        if _is_spreadsheet_url(href) or "ranking" in _strip_accents(text).lower():
            return href
    return None


def _fetch_url(url: str):
    try:
        response = federation_request(
            "get",
            url,
            headers=HEADERS,
            timeout=30,
            allow_redirects=True,
        )
    except requests.RequestException as exc:
        print(f"    Request error for {url}: {exc}")
        return None

    if response.status_code != 200:
        print(f"    HTTP {response.status_code} for {url}")
        return None
    return response


def _download_category_workbook(category: str) -> bytes | None:
    if category in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[category]

    url = _discover_category_links().get(category)
    if not url:
        print(f"    No public Puerto Rico ranking link for {category}")
        return None

    response = _fetch_url(url)
    if response is None:
        return None

    final_url = getattr(response, "url", url) or url
    if _spreadsheet_response(response, final_url):
        _WORKBOOK_CACHE[category] = response.content
        return response.content

    if _looks_unavailable(response.text):
        print(f"    No scrapeable rankings at {final_url}: blocked/login/JS-only page")
        return None

    spreadsheet_url = _first_spreadsheet_link(response.text, base_url=final_url)
    if not spreadsheet_url:
        print(f"    No spreadsheet ranking link found at {final_url}")
        return None

    spreadsheet_response = _fetch_url(spreadsheet_url)
    if spreadsheet_response is None:
        return None
    final_spreadsheet_url = getattr(spreadsheet_response, "url", spreadsheet_url) or spreadsheet_url
    if not _spreadsheet_response(spreadsheet_response, final_spreadsheet_url):
        print(f"    Ranking link is not a spreadsheet: {final_spreadsheet_url}")
        return None

    _WORKBOOK_CACHE[category] = spreadsheet_response.content
    return spreadsheet_response.content


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _worksheet_text(worksheet) -> str:
    lines = []
    for row in worksheet.iter_rows(values_only=True):
        cells = [_format_cell(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        if cells and any(cell for cell in cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _combo_words(weapon: str, gender: str, category: str) -> tuple[set[str], set[str], set[str]]:
    return (
        {_token(word) for word in _WEAPON_ALIASES[weapon]},
        {_token(word) for word in _GENDER_ALIASES[gender]},
        {_token(word) for word in _CATEGORY_ALIASES[category]},
    )


def _worksheet_matches_combo(worksheet, weapon: str, gender: str, category: str) -> bool:
    context_parts = [worksheet.title]
    for index, row in enumerate(worksheet.iter_rows(values_only=True)):
        if index >= 10:
            break
        values = [_clean_text(value) for value in row if _clean_text(value)]
        if values:
            context_parts.append(" ".join(values))

    words = {_token(word) for word in _normalised_words(" ".join(context_parts))}
    weapon_words, gender_words, category_words = _combo_words(weapon, gender, category)
    weapon_match = bool(words.intersection(weapon_words))
    gender_match = bool(words.intersection(gender_words))
    category_match = bool(words.intersection(category_words))
    return weapon_match and gender_match and category_match


def _extract_xlsx_combo_text(content: bytes, weapon: str, gender: str, category: str) -> str | None:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        print(f"    XLSX open error: {exc}")
        return None

    for worksheet in workbook.worksheets:
        if _worksheet_matches_combo(worksheet, weapon, gender, category):
            return _worksheet_text(worksheet)

    return None


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Fetch one Puerto Rico ranking combo, returning None on 404/network/missing data."""
    content = _download_category_workbook(category)
    if content is None:
        return None

    text = _extract_xlsx_combo_text(content, weapon, gender, category)
    if text is None:
        print(f"    No public sheet for {weapon} {gender} {category}")
    return text


def _combo_label(combo: tuple[str, str, str]) -> str:
    weapon, gender, category = combo
    return f"{weapon} {gender} {category}"


def main():
    run_log = ScraperRunLogger("scrape_fed_pur").start()
    season = current_season()
    print(f"Puerto Rico federation rankings - season {season}")

    total_written = 0
    total_failed = 0
    total_skipped = 0
    working_combos: list[str] = []
    failed_combos: list[str] = []

    try:
        for combo in RANKING_COMBOS:
            weapon, gender, category = combo
            label = _combo_label(combo)
            print(f"  {label}...")

            content = fetch_rankings_page(weapon, gender, category)
            if content is None:
                failed_combos.append(label)
                total_failed += 1
                time.sleep(REQUEST_DELAY)
                continue

            parsed = parse_rankings_table(content)
            if not parsed:
                print("    No rows parsed")
                failed_combos.append(label)
                total_failed += 1
                time.sleep(REQUEST_DELAY)
                continue

            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                    metadata={
                        "country_code": COUNTRY_CODE,
                        "ranking_page": RANKING_PAGE,
                        "data_format": "xlsx",
                    },
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            print(f"    Parsed {len(parsed)} rows; written {written}")
            total_written += written
            working_combos.append(label)
            time.sleep(REQUEST_DELAY)

        metadata = {
            "season": season,
            "combos_working": len(working_combos),
            "combos_total": len(RANKING_COMBOS),
            "working_combos": working_combos,
            "failed_combos": failed_combos,
            "ranking_page": RANKING_PAGE,
            "data_format": "xlsx",
        }
        set_state(SOURCE, "last_run", metadata)
        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata=metadata,
        )
        print(
            f"Done - written={total_written}, failed={total_failed}, skipped={total_skipped}, "
            f"combos_working={len(working_combos)}/{len(RANKING_COMBOS)}"
        )
        if failed_combos:
            print(f"Failed combos: {', '.join(failed_combos)}")
    except Exception as exc:
        set_state(SOURCE, "last_error", {"season": season, "error": str(exc)})
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
