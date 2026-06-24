"""
scrape_fed_irl.py - Fencing Ireland national rankings scraper.

Probe notes, 2026-06-02:
  - The requested irishfencing.net host did not resolve from the local probe.
  - Current public federation pages are under https://www.fencingireland.net/.
  - The federation menu links Senior Rankings to a public Google Sheet:
    https://docs.google.com/spreadsheets/d/1iZdJ_GfFRx61_qwvYa5Ck9dTKN3lM852zfDSf2Cvw-g/edit
  - Method: GET.
  - Response format: public Google Sheets HTML for browser view; scraper uses
    the public XLSX export and extracts one worksheet per weapon/gender/category.
  - Visible headers: Rank | Fencer | Club | Points.
  - The public cadet/junior page did not expose durable Junior ranking links
    in the rendered probe. Junior combos are still attempted and fail closed
    if no matching worksheet exists.
"""

from __future__ import annotations

import io
import re
import time
import unicodedata
from datetime import UTC, datetime, timezone

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import get_state, set_state

SOURCE = "irl_fencing"
COUNTRY = "Ireland"
BASE_URL = "https://www.fencingireland.net"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    ),
    "Accept-Language": "en-IE,en;q=0.9",
    "Referer": f"{BASE_URL}/",
}

SENIOR_RANKINGS_SHEET_ID = "1iZdJ_GfFRx61_qwvYa5Ck9dTKN3lM852zfDSf2Cvw-g"
SENIOR_RANKINGS_URL = (
    f"https://docs.google.com/spreadsheets/d/{SENIOR_RANKINGS_SHEET_ID}/edit"
)
SENIOR_RANKINGS_XLSX_URL = (
    f"https://docs.google.com/spreadsheets/d/{SENIOR_RANKINGS_SHEET_ID}/export?format=xlsx"
)
DATA_FORMAT = "xlsx"

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

_WORKBOOK_CACHE: dict[str, bytes] = {}

_RANK_HEADERS = {"#", "rank", "rang", "place", "position", "pos"}
_NAME_HEADERS = {"name", "fencer", "athlete", "nom", "fullname", "competitor"}
_CLUB_HEADERS = {"club", "clubs", "team", "school", "affiliation"}
_POINT_HEADERS = {"points", "point", "pts", "score", "totalpoints", "rankingpoints"}
_SKIP_TOKENS = {
    "",
    "dns",
    "dnf",
    "dq",
    "dsq",
    "wd",
    "withdrawn",
    "disqualified",
    "total",
    "totals",
    "summary",
    "subtotal",
    "no data",
}
_NO_DATA_MARKERS = {
    "no rankings available",
    "no ranking available",
    "no data",
    "please select another ranking file",
}
_BLOCKED_MARKERS = {
    "javascript isn't enabled",
    "sign in",
    "accounts.google.com",
    "login",
    "access denied",
    "forbidden",
}


def _clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def _strip_accents(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _header_key(value: str) -> str:
    text = _strip_accents(_clean_text(value).lower())
    if text == "#":
        return "#"
    text = text.replace("&", " and ").replace("/", " ")
    return re.sub(r"[^a-z0-9]+", "", text)


def _label_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", _strip_accents(_clean_text(value).lower()))


def _parse_rank(value: str) -> int | None:
    text = _clean_text(value).rstrip(".")
    if _header_key(text) in _SKIP_TOKENS:
        return None
    match = re.fullmatch(r"(\d+)(?:\.0+)?", text)
    if not match:
        return None
    rank = int(match.group(1))
    return rank if rank > 0 else None


def _parse_points(value: str) -> float | None:
    text = _clean_text(value)
    if not text or _header_key(text) in _SKIP_TOKENS or text in {"-", "--", "—", "–"}:
        return None

    number = text.replace("\u00a0", "").replace(" ", "").replace("'", "")
    number = re.sub(r"[^0-9,.\-]", "", number)
    if not number or number in {"-", ".", ","}:
        return None

    if "," in number and "." in number:
        if number.rfind(",") > number.rfind("."):
            number = number.replace(".", "").replace(",", ".")
        else:
            number = number.replace(",", "")
    elif "," in number:
        if re.fullmatch(r"-?\d{1,3}(,\d{3})+", number):
            number = number.replace(",", "")
        else:
            number = number.replace(",", ".")
    elif "." in number and re.fullmatch(r"-?\d{1,3}(\.\d{3})+", number):
        number = number.replace(".", "")

    try:
        return float(number)
    except ValueError:
        return None


def _detect_columns(cells: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    point_candidates: list[tuple[int, int]] = []

    for index, cell in enumerate(cells):
        key = _header_key(cell)
        if key in _RANK_HEADERS and "rank_col" not in mapping:
            mapping["rank_col"] = index
        elif key in _NAME_HEADERS and "name_col" not in mapping:
            mapping["name_col"] = index
        elif key in _CLUB_HEADERS and "club_col" not in mapping:
            mapping["club_col"] = index
        elif key in _POINT_HEADERS or key.endswith("points"):
            score = 0 if key in {"points", "pts", "totalpoints"} else 1
            point_candidates.append((score, index))

    if point_candidates:
        mapping["points_col"] = min(point_candidates)[1]
    return mapping


def _append_parsed_row(rows: list[dict], columns: dict[str, int], cells: list[str]) -> None:
    if "rank_col" not in columns or "name_col" not in columns:
        return
    required_max = max(columns["rank_col"], columns["name_col"])
    if len(cells) <= required_max:
        return

    rank = _parse_rank(cells[columns["rank_col"]])
    if rank is None:
        return

    name = _clean_text(cells[columns["name_col"]])
    if not name or _header_key(name) in _SKIP_TOKENS:
        return

    club = None
    club_col = columns.get("club_col")
    if club_col is not None and club_col < len(cells):
        club = _clean_text(cells[club_col]) or None

    points = None
    points_col = columns.get("points_col")
    if points_col is not None and points_col < len(cells):
        points = _parse_points(cells[points_col])

    rows.append({"rank": rank, "name": name, "club": club, "points": points})


def _rows_from_matrix(matrix: list[list[str]]) -> list[dict]:
    columns: dict[str, int] | None = None
    parsed: list[dict] = []

    for cells in matrix:
        cells = [_clean_text(cell) for cell in cells]
        if not any(cells):
            continue

        candidate = _detect_columns(cells)
        if "rank_col" in candidate and "name_col" in candidate:
            columns = candidate
            continue

        if columns is None:
            rank = _parse_rank(cells[0]) if cells else None
            if rank is None or len(cells) < 2:
                continue
            columns = {"rank_col": 0, "name_col": 1}
            if len(cells) > 2:
                columns["club_col"] = 2
            if len(cells) > 3:
                columns["points_col"] = 3

        _append_parsed_row(parsed, columns, cells)

    return parsed


def _html_matrix(html: str) -> list[list[str]]:
    soup = BeautifulSoup(html, "html.parser")
    best: list[list[str]] = []
    for table in soup.find_all("table"):
        matrix: list[list[str]] = []
        for row in table.find_all("tr"):
            if row.find_parent("table") is not table:
                continue
            cells = row.find_all(["td", "th"], recursive=False)
            if cells:
                matrix.append([_clean_text(cell.get_text(" ", strip=True)) for cell in cells])
        if len(matrix) > len(best):
            best = matrix
    return best


def _text_matrix(text: str) -> list[list[str]]:
    matrix: list[list[str]] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if "\t" in line:
            cells = line.split("\t")
        elif "|" in line:
            cells = line.split("|")
        else:
            cells = re.split(r"\s{2,}", line)
        matrix.append([_clean_text(cell) for cell in cells])
    return matrix


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Ireland ranking rows into rank/name/club/points dictionaries."""
    if not html_or_text or not html_or_text.strip():
        return []

    lowered = html_or_text.lower()
    if any(marker in lowered for marker in _NO_DATA_MARKERS):
        return []

    if "<table" in lowered:
        rows = _rows_from_matrix(_html_matrix(html_or_text))
        if rows:
            return rows

    return _rows_from_matrix(_text_matrix(html_or_text))


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _sheet_candidates(weapon: str, gender: str, category: str) -> set[str]:
    weapon_variants = {
        "Foil": ["foil", "floret", "fleuret"],
        "Epee": ["epee", "epée", "degen"],
        "Sabre": ["sabre", "saber", "sabel"],
    }.get(weapon, [weapon])
    gender_variants = {
        "Men": ["men", "mens", "men's", "male", "m"],
        "Women": ["women", "womens", "women's", "female", "w"],
    }.get(gender, [gender])
    category_variants = {
        "Senior": ["senior", "seniors", "open", "adult"],
        "Junior": ["junior", "juniors", "u20", "u-20", "under20"],
    }.get(category, [category])

    candidates: set[str] = set()
    for weapon_name in weapon_variants:
        for gender_name in gender_variants:
            candidates.add(_label_key(f"{gender_name} {weapon_name}"))
            candidates.add(_label_key(f"{weapon_name} {gender_name}"))
            for category_name in category_variants:
                candidates.add(_label_key(f"{category_name} {gender_name} {weapon_name}"))
                candidates.add(_label_key(f"{category_name} {weapon_name} {gender_name}"))

    category_code = "S" if category == "Senior" else "J"
    gender_code = "M" if gender == "Men" else "W"
    weapon_code = {"Foil": "F", "Epee": "E", "Sabre": "S"}[weapon]
    candidates.add(_label_key(f"{category_code}{gender_code}{weapon_code}"))
    candidates.add(_label_key(f"{gender_code}{weapon_code}"))
    return candidates


def _worksheet_matches(title: str, weapon: str, gender: str, category: str) -> bool:
    key = _label_key(title)
    if key in _sheet_candidates(weapon, gender, category):
        return True

    has_weapon = {
        "Foil": any(token in key for token in ("foil", "floret", "fleuret")),
        "Epee": any(token in key for token in ("epee", "degen")),
        "Sabre": any(token in key for token in ("sabre", "saber", "sabel")),
    }[weapon]
    has_gender = {
        "Men": "men" in key and "women" not in key,
        "Women": "women" in key or key.startswith("w"),
    }[gender]
    if not has_weapon or not has_gender:
        return False

    if category == "Junior":
        return any(token in key for token in ("junior", "juniors", "u20", "under20"))
    return "junior" not in key and "u20" not in key


def _extract_xlsx_sheet_text(content: bytes, weapon: str, gender: str, category: str) -> str | None:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        print(f"    XLSX open error: {exc}")
        return None

    worksheet = None
    for candidate in workbook.worksheets:
        if _worksheet_matches(candidate.title, weapon, gender, category):
            worksheet = candidate
            break

    if worksheet is None:
        print(f"    No Ireland worksheet for {weapon} {gender} {category}")
        return None

    lines: list[str] = []
    for row in worksheet.iter_rows(values_only=True):
        cells = [_format_cell(value) for value in row]
        while cells and not cells[-1]:
            cells.pop()
        if cells and any(cells):
            lines.append("\t".join(cells))
    return "\n".join(lines)


def _looks_blocked_or_js_only(response) -> bool:
    content_type = response.headers.get("content-type", "").lower()
    if "html" not in content_type and not response.content.lstrip().startswith(b"<"):
        return False
    text = (response.text or response.content[:2000].decode("utf-8", errors="ignore")).lower()
    return any(marker in text for marker in _BLOCKED_MARKERS)


def _download_workbook() -> bytes | None:
    if SENIOR_RANKINGS_XLSX_URL in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[SENIOR_RANKINGS_XLSX_URL]

    try:
        response = federation_request(
            "get",
            SENIOR_RANKINGS_XLSX_URL,
            headers=HEADERS,
            timeout=45,
            allow_redirects=True,
        )
    except requests.RequestException as exc:
        print(f"    Request error for {SENIOR_RANKINGS_XLSX_URL}: {exc}")
        return None

    if response.status_code != 200:
        print(f"    HTTP {response.status_code} for {SENIOR_RANKINGS_XLSX_URL}")
        return None

    if _looks_blocked_or_js_only(response):
        print(f"    No scrapeable rankings at {SENIOR_RANKINGS_XLSX_URL}")
        return None

    content_type = response.headers.get("content-type", "").lower()
    if "spreadsheetml.sheet" not in content_type and not response.content.startswith(b"PK\x03\x04"):
        print(f"    Unexpected Ireland ranking format: {content_type or 'unknown'}")
        return None

    _WORKBOOK_CACHE[SENIOR_RANKINGS_XLSX_URL] = response.content
    return response.content


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Fetch ranking content for one Ireland weapon/gender/category combo."""
    workbook = _download_workbook()
    if not workbook:
        return None
    return _extract_xlsx_sheet_text(workbook, weapon, gender, category)


def current_season() -> str:
    """Return the current fencing season as YYYY-YYYY."""
    now = datetime.now(UTC)
    start_year = now.year - 1 if now.month < 7 else now.year
    season = f"{start_year}-{start_year + 1}"
    try:
        import season_utils  # type: ignore

        if hasattr(season_utils, "normalize_season"):
            normalized = season_utils.normalize_season(season)
            if isinstance(normalized, str):
                return normalized
    except Exception:
        pass
    return season


def _row_metadata(weapon: str, gender: str, category: str) -> dict:
    return {
        "source_url": SENIOR_RANKINGS_URL,
        "file_url": SENIOR_RANKINGS_XLSX_URL,
        "official_federation_url": BASE_URL,
        "data_format": DATA_FORMAT,
        "source_language": "en",
        "combo": f"{weapon} {gender} {category}",
    }


def main() -> None:
    run_log = ScraperRunLogger("scrape_fed_irl").start()
    season = current_season()
    previous_state = get_state(SOURCE, "last_run")
    if previous_state:
        print(f"Previous Ireland federation run state found: {previous_state}")

    print(f"Fencing Ireland rankings - season {season}")
    print(f"Ranking source: {SENIOR_RANKINGS_URL}")

    total_written = 0
    total_failed = 0
    total_skipped = 0
    working_combos = 0
    failed_combos: list[str] = []
    skipped_combos: list[str] = []

    try:
        for index, (weapon, gender, category) in enumerate(RANKING_COMBOS):
            combo_label = f"{weapon} {gender} {category}"
            print(f"  {combo_label}...")

            content = fetch_rankings_page(weapon, gender, category)
            if not content:
                if category == "Junior":
                    total_skipped += 1
                    skipped_combos.append(combo_label)
                else:
                    total_failed += 1
                    failed_combos.append(combo_label)
                if index < len(RANKING_COMBOS) - 1:
                    time.sleep(REQUEST_DELAY)
                continue

            parsed = parse_rankings_table(content)
            if not parsed:
                print(f"    No rows parsed for {combo_label}")
                if category == "Junior":
                    total_skipped += 1
                    skipped_combos.append(combo_label)
                else:
                    total_failed += 1
                    failed_combos.append(combo_label)
                if index < len(RANKING_COMBOS) - 1:
                    time.sleep(REQUEST_DELAY)
                continue

            metadata = _row_metadata(weapon, gender, category)
            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                    metadata=metadata,
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            print(f"    Parsed {len(rows)} rows; written {written}")
            total_written += written
            working_combos += 1

            if index < len(RANKING_COMBOS) - 1:
                time.sleep(REQUEST_DELAY)

        state = {
            "season": season,
            "working_combos": working_combos,
            "attempted_combos": len(RANKING_COMBOS),
            "failed_combos": failed_combos,
            "skipped_combos": skipped_combos,
            "source_url": SENIOR_RANKINGS_URL,
            "data_format": DATA_FORMAT,
            "updated_at": datetime.now(UTC).isoformat(),
        }
        set_state(SOURCE, "last_run", state)

        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata=state,
        )
        if failed_combos:
            print(f"Failed combos: {', '.join(failed_combos)}")
        if skipped_combos:
            print(f"Skipped combos: {', '.join(skipped_combos)}")
        print(
            "Done - "
            f"written={total_written}, failed={total_failed}, skipped={total_skipped}, "
            f"working_combos={working_combos}/12"
        )
    except Exception as exc:
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
