"""
scrape_fed_tpe.py - Chinese Taipei federation rankings scraper.

Probe evidence, 2026-06-02:
  - Official site: https://www.fencing.org.tw/
  - Request method: GET
  - Response format: public HTML index plus public XLSX files hosted on x.webdo.cc.
  - Current homepage has a "Ranking 選手排名" section with current youth files:
      青年組排名(115-1)(公告版).xlsx
      青少年組排名(115-1)(公告版).xlsx
      少年組排名(115-1)(公告版).xlsx
  - Public Junior source: 青年組 ranking workbook, covering Foil/Epee/Sabre Men/Women.
  - Current public Senior full-ranking workbook was not visible in the probed homepage.

Traditional Chinese weapon labels:
  銳劍 = Epee, 鈍劍 = Foil, 軍刀 = Sabre
"""

from __future__ import annotations

import io
import re
import time
import unicodedata
from datetime import datetime, timezone
from urllib.parse import unquote, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import set_state

try:
    from season_utils import season_to_string
except ImportError:  # pragma: no cover - compatibility fallback for older checkouts
    def season_to_string(season_int: int) -> str:
        return f"{season_int - 1:04d}-{season_int:04d}"


SOURCE = "tpe_fencing"
COUNTRY = "Chinese Taipei"
BASE_URL = "https://www.fencing.org.tw/"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
        "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8"
    ),
    "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
}

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

RANKING_INDEX_URLS = [
    BASE_URL,
    (
        "https://www.fencing.org.tw/portal_c3_cnt.php?"
        "button_num=c3&folder_id=99&owner_num=c3_6578&search_type=1&search_word=%E6%8E%92%E5%90%8D"
    ),
    (
        "https://www.fencing.org.tw/portal_c3_cnt.php?"
        "button_num=c3&folder_id=34&owner_num=c3_4936&search_type=1&search_word=%E6%8E%92%E5%90%8D"
    ),
]

_RANKING_FILE_CACHE: dict[str, str] = {}
_WORKBOOK_CACHE: dict[str, bytes] = {}

_RANK_HEADERS = {"#", "rank", "ranking", "place", "position"}
_NAME_HEADERS = {"name", "fencer", "athlete"}
_CLUB_HEADERS = {"club", "clubs", "school", "team", "unit", "organization"}
_POINT_HEADERS = {"points", "pts", "score", "totalpoints"}
_SKIP_TOKENS = {
    "",
    "dns",
    "dq",
    "dsq",
    "dnf",
    "wd",
    "ret",
    "total",
    "summary",
    "subtotal",
    "rank",
    "ranking",
    "name",
    "合計",
    "總計",
    "总计",
    "小計",
    "小计",
    "摘要",
    "備註",
    "备注",
    "名次",
    "排名",
    "姓名",
}
_NO_DATA_MARKERS = {
    "目前無公開排名資料",
    "無資料",
    "查無資料",
    "沒有資料",
    "no rankings available",
    "no ranking available",
    "no data",
}
_LOGIN_OR_BLOCKED_MARKERS = {
    "會員登入",
    "会员登入",
    "請先登入",
    "请先登入",
    "login required",
    "access denied",
    "forbidden",
    "captcha",
}
_JS_ONLY_MARKERS = {
    "please enable javascript",
    "enable javascript",
    "請開啟javascript",
    "请开启javascript",
}

_WEAPON_LABELS = {
    "Epee": ("銳劍", "锐剑", "銳", "锐", "epee", "épée", "重劍", "重剑"),
    "Foil": ("鈍劍", "钝剑", "鈍", "钝", "foil", "花劍", "花剑"),
    "Sabre": ("軍刀", "军刀", "軍", "军", "sabre", "saber", "佩劍", "佩剑"),
}
_GENDER_LABELS = {
    "Men": ("男子", "男", "men", "male"),
    "Women": ("女子", "女", "women", "female"),
}


def current_season() -> str:
    """Return the current federation season range as YYYY-YYYY."""
    now = datetime.now(timezone.utc)
    season_end_year = now.year if now.month < 7 else now.year + 1
    return season_to_string(season_end_year)


def _compact_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def _ascii_key(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    ascii_text = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r"[^a-z0-9#]+", "", ascii_text.lower())


def _is_rank_header(value: str) -> bool:
    text = _compact_text(value)
    return text in {"名次", "排名", "#"} or _ascii_key(text) in _RANK_HEADERS


def _is_name_header(value: str) -> bool:
    text = _compact_text(value)
    return "姓名" in text or "選手" in text or "选手" in text or _ascii_key(text) in _NAME_HEADERS


def _is_club_header(value: str) -> bool:
    text = _compact_text(value)
    return (
        "單位" in text
        or "单位" in text
        or "俱樂部" in text
        or "俱乐部" in text
        or "學校" in text
        or "学校" in text
        or _ascii_key(text) in _CLUB_HEADERS
    )


def _is_points_header(value: str) -> bool:
    text = _compact_text(value)
    return "積分" in text or "积分" in text or _ascii_key(text) in _POINT_HEADERS


def _is_skip_text(value: str) -> bool:
    text = _compact_text(value)
    key = _ascii_key(text)
    return text in _SKIP_TOKENS or (bool(key) and key in _SKIP_TOKENS)


def _parse_rank(value: str) -> int | None:
    text = _compact_text(value).rstrip(".")
    if not text or _is_skip_text(text):
        return None
    match = re.match(r"^\s*(\d+)", text)
    if not match:
        return None
    rank = int(match.group(1))
    return rank if rank > 0 else None


def _parse_points(value: str) -> float | None:
    text = _compact_text(value)
    if not text or _is_skip_text(text) or text in {"-", "—", "–"}:
        return None

    text = re.sub(r"[^0-9,.\-]", "", text)
    if text in {"", "-", ".", ","}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        if len(parts) > 2:
            text = "".join(parts[:-1]) + "." + parts[-1]
        else:
            left, right = parts
            if len(right) == 3 and left.lstrip("-").isdigit():
                text = left + right
            else:
                text = left + "." + right
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 2 and all(len(part) == 3 for part in parts[1:]):
            text = "".join(parts)

    try:
        return float(text)
    except ValueError:
        return None


def _html_tables_to_matrices(html: str) -> list[list[list[str]]]:
    soup = BeautifulSoup(html, "html.parser")
    matrices = []
    for table in soup.find_all("table"):
        matrix = []
        for row in table.find_all("tr"):
            cells = [_compact_text(cell.get_text(" ", strip=True)) for cell in row.find_all(["td", "th"])]
            if cells:
                matrix.append(cells)
        if matrix:
            matrices.append(matrix)
    return matrices


def _text_to_matrix(text: str) -> list[list[str]]:
    matrix = []
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        if "\t" in line:
            cells = line.split("\t")
        elif "|" in line:
            cells = line.split("|")
        else:
            cells = re.split(r"\s{2,}", line)
        cells = [_compact_text(cell) for cell in cells]
        if len(cells) > 1:
            matrix.append(cells)
    return matrix


def _detect_columns(cells: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell in enumerate(cells):
        if "rank" not in mapping and _is_rank_header(cell):
            mapping["rank"] = idx
        elif "name" not in mapping and _is_name_header(cell):
            mapping["name"] = idx
        elif "club" not in mapping and _is_club_header(cell):
            mapping["club"] = idx
        elif "points" not in mapping and _is_points_header(cell):
            mapping["points"] = idx
    return mapping


def _parse_matrix(matrix: list[list[str]]) -> list[dict]:
    rows: list[dict] = []
    seen: set[tuple[int, str]] = set()

    for header_idx, header in enumerate(matrix):
        col_map = _detect_columns(header)
        if "rank" not in col_map or "name" not in col_map:
            continue

        required_max = max(col_map.values())
        for cells in matrix[header_idx + 1:]:
            if len(cells) <= required_max:
                continue

            rank = _parse_rank(cells[col_map["rank"]])
            if rank is None:
                continue

            name = _compact_text(cells[col_map["name"]])
            if not name or _is_skip_text(name):
                continue

            club = None
            club_idx = col_map.get("club")
            if club_idx is not None and club_idx < len(cells):
                club = _compact_text(cells[club_idx]) or None

            points = None
            points_idx = col_map.get("points")
            if points_idx is not None and points_idx < len(cells):
                points = _parse_points(cells[points_idx])

            key = (rank, name)
            if key in seen:
                continue
            seen.add(key)
            rows.append({"rank": rank, "name": name, "club": club, "points": points})
        break

    return rows


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Chinese Taipei ranking content into rank/name/club/points rows."""
    if not html_or_text or not html_or_text.strip():
        return []

    lowered = html_or_text.lower()
    if any(marker in html_or_text for marker in _NO_DATA_MARKERS) or any(marker in lowered for marker in _NO_DATA_MARKERS):
        return []

    parsed_rows: list[dict] = []
    for matrix in _html_tables_to_matrices(html_or_text):
        parsed_rows.extend(_parse_matrix(matrix))
    if parsed_rows:
        return parsed_rows

    return _parse_matrix(_text_to_matrix(BeautifulSoup(html_or_text, "html.parser").get_text("\n", strip=True)))


def _response_text(response) -> str:
    text = getattr(response, "text", "")
    if text:
        return text
    content = getattr(response, "content", b"") or b""
    try:
        return content.decode("utf-8", errors="ignore")
    except AttributeError:
        return str(content)


def _is_unusable_page(response) -> bool:
    status_code = getattr(response, "status_code", 0)
    if status_code in {401, 403}:
        return True

    text = _response_text(response)
    lowered = text.lower()
    return (
        any(marker in text for marker in _LOGIN_OR_BLOCKED_MARKERS)
        or any(marker in lowered for marker in _LOGIN_OR_BLOCKED_MARKERS)
        or any(marker in lowered for marker in _JS_ONLY_MARKERS)
    )


def _category_from_link(text: str, href: str) -> str | None:
    combined = f"{text} {unquote(href)}"
    lowered = combined.lower()
    if not lowered.endswith((".xlsx", ".xls")) and not any(ext in lowered for ext in (".xlsx", ".xls")):
        return None

    if "青年組排名" in combined or "青年组排名" in combined or "junior" in lowered:
        return "Junior"
    if (
        "全國排名" in combined
        or "全国排名" in combined
        or "全國最新排名" in combined
        or "全國積分排名" in combined
        or "senior" in lowered
    ):
        if "青年" not in combined and "青少年" not in combined and "少年" not in combined:
            return "Senior"
    return None


def _extract_ranking_file_links(html: str, *, base_url: str = BASE_URL) -> dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")
    links: dict[str, str] = {}
    for anchor in soup.find_all("a", href=True):
        href = urljoin(base_url, anchor["href"])
        text = _compact_text(anchor.get_text(" ", strip=True))
        category = _category_from_link(text, href)
        if category and category not in links:
            links[category] = href
    return links


def _discover_ranking_files() -> dict[str, str]:
    if _RANKING_FILE_CACHE:
        return dict(_RANKING_FILE_CACHE)

    discovered: dict[str, str] = {}
    for url in RANKING_INDEX_URLS:
        try:
            response = federation_request("get", url, headers=HEADERS, timeout=25, allow_redirects=True)
        except requests.RequestException as exc:
            print(f"    Ranking index request error for {url}: {exc}")
            continue

        if getattr(response, "status_code", 0) != 200:
            print(f"    Ranking index HTTP {response.status_code} for {url}")
            continue
        if _is_unusable_page(response):
            print(f"    Ranking index unavailable or login/JS-only at {url}")
            continue

        for category, href in _extract_ranking_file_links(_response_text(response), base_url=getattr(response, "url", url)).items():
            discovered.setdefault(category, href)

    _RANKING_FILE_CACHE.update(discovered)
    return dict(discovered)


def _download_workbook(url: str) -> bytes | None:
    if url in _WORKBOOK_CACHE:
        return _WORKBOOK_CACHE[url]

    try:
        response = federation_request("get", url, headers=HEADERS, timeout=30, allow_redirects=True)
    except requests.RequestException as exc:
        print(f"    Workbook request error for {url}: {exc}")
        return None

    if getattr(response, "status_code", 0) != 200:
        print(f"    Workbook HTTP {response.status_code} for {url}")
        return None
    if _is_unusable_page(response):
        print(f"    Workbook unavailable or login/JS-only at {url}")
        return None

    content = getattr(response, "content", b"") or b""
    if not content:
        return None

    _WORKBOOK_CACHE[url] = bytes(content)
    return _WORKBOOK_CACHE[url]


def _matches_combo(text: str, weapon: str, gender: str) -> bool:
    lowered = text.lower()
    return any(label in text or label in lowered for label in _WEAPON_LABELS[weapon]) and any(
        label in text or label in lowered for label in _GENDER_LABELS[gender]
    )


def _row_slice(matrix: list[list[str]], row_idx: int, start: int, end: int) -> list[str]:
    if row_idx >= len(matrix):
        return []
    row = matrix[row_idx]
    return [row[i] if i < len(row) else "" for i in range(start, end)]


def _group_label(matrix: list[list[str]], header_idx: int, start: int, end: int) -> str:
    parts: list[str] = []
    for row_idx in range(max(0, header_idx - 3), header_idx + 1):
        parts.extend(_row_slice(matrix, row_idx, start, end))
    return " ".join(part for part in parts if part)


def _group_columns(header: list[str], start: int, end: int) -> dict[str, int]:
    group = [header[i] if i < len(header) else "" for i in range(start, end)]
    col_map = _detect_columns(group)
    if "rank" in col_map:
        col_map["rank"] += start
    else:
        col_map["rank"] = start
    if "name" in col_map:
        col_map["name"] += start
    elif start + 1 < end:
        col_map["name"] = start + 1
    if "club" in col_map:
        col_map["club"] += start
    if "points" in col_map:
        col_map["points"] += start
    return col_map


def _extract_combo_from_matrix(matrix: list[list[str]], weapon: str, gender: str) -> str | None:
    for header_idx, header in enumerate(matrix):
        rank_starts = [idx for idx, cell in enumerate(header) if _is_rank_header(cell)]
        if not rank_starts:
            continue

        for pos, start in enumerate(rank_starts):
            end = rank_starts[pos + 1] if pos + 1 < len(rank_starts) else len(header)
            if not _matches_combo(_group_label(matrix, header_idx, start, end), weapon, gender):
                continue

            col_map = _group_columns(header, start, end)
            if "rank" not in col_map or "name" not in col_map:
                continue

            rows = ["名次\t姓名\t單位\t積分"]
            required_max = max(col_map.values())
            for row in matrix[header_idx + 1:]:
                if len(row) <= required_max:
                    continue
                rank = row[col_map["rank"]]
                name = row[col_map["name"]]
                if _parse_rank(rank) is None or not _compact_text(name) or _is_skip_text(name):
                    continue

                club = ""
                club_idx = col_map.get("club")
                if club_idx is not None and club_idx < len(row):
                    club = row[club_idx]

                points = ""
                points_idx = col_map.get("points")
                if points_idx is not None and points_idx < len(row):
                    points = row[points_idx]

                rows.append("\t".join([_compact_text(rank), _compact_text(name), _compact_text(club), _compact_text(points)]))

            if len(rows) > 1:
                return "\n".join(rows)

    return None


def _workbook_combo_to_text(content: bytes, weapon: str, gender: str, category: str) -> str | None:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        matrix: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            cells = [_compact_text(cell) for cell in row]
            while cells and not cells[-1]:
                cells.pop()
            if cells:
                matrix.append(cells)

        if not matrix:
            continue

        combo_text = _extract_combo_from_matrix(matrix, weapon, gender)
        if combo_text:
            return combo_text

        sheet_context = f"{worksheet.title} " + " ".join(" ".join(row) for row in matrix[:4])
        if _matches_combo(sheet_context, weapon, gender):
            text = "\n".join("\t".join(row) for row in matrix)
            if parse_rankings_table(text):
                return text

    print(f"    No {weapon} {gender} section found in {category} workbook")
    return None


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Fetch and normalize one Chinese Taipei ranking combo. Returns None on failures."""
    files = _discover_ranking_files()
    url = files.get(category)
    if not url:
        print(f"    No scrapeable rankings at {BASE_URL} for {weapon} {gender} {category}")
        return None

    content = _download_workbook(url)
    if not content:
        return None

    try:
        return _workbook_combo_to_text(content, weapon, gender, category)
    except Exception as exc:
        print(f"    Workbook parse error for {url}: {exc}")
        return None


def _combo_label(weapon: str, gender: str, category: str) -> str:
    return f"{weapon} {gender} {category}"


def main():
    run_log = ScraperRunLogger("scrape_fed_tpe").start()
    season = current_season()
    print(f"Chinese Taipei federation rankings - season {season}")

    total_written = 0
    total_failed = 0
    total_skipped = 0
    combos_working = 0
    failed_combos: list[str] = []

    try:
        for weapon, gender, category in RANKING_COMBOS:
            label = _combo_label(weapon, gender, category)
            print(f"  {label}...")

            content = fetch_rankings_page(weapon, gender, category)
            if not content:
                total_failed += 1
                failed_combos.append(f"{label}: No scrapeable rankings at {BASE_URL}")
                time.sleep(REQUEST_DELAY)
                continue

            parsed = parse_rankings_table(content)
            if not parsed:
                print("    No rows parsed")
                total_failed += 1
                failed_combos.append(f"{label}: no rows parsed")
                time.sleep(REQUEST_DELAY)
                continue

            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                    metadata={"source_url": BASE_URL, "source_format": "html+xlsx"},
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            print(f"    Written {written} rows ({len(parsed)} parsed)")
            total_written += written
            combos_working += 1
            time.sleep(REQUEST_DELAY)

        summary = {
            "season": season,
            "combos": len(RANKING_COMBOS),
            "combos_working": combos_working,
            "failed_combos": failed_combos,
            "probe": {
                "working_url": BASE_URL,
                "method": "GET",
                "response_format": "public HTML index plus XLSX workbook downloads",
                "public_combos": "Junior Foil/Epee/Sabre Men/Women; Senior not found in current public homepage probe",
            },
        }
        set_state(SOURCE, "last_run", summary)
        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata=summary,
        )
        print(
            "Done - "
            f"written={total_written}, failed={total_failed}, skipped={total_skipped}, "
            f"combos_working={combos_working}/{len(RANKING_COMBOS)}"
        )
    except Exception as exc:
        set_state(SOURCE, "last_error", {"season": season, "error": str(exc)})
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
