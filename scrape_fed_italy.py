"""
scrape_fed_italy.py — Italian Fencing Federation national rankings scraper.

Federscherma publishes ranking documents as spreadsheet downloads. Current
documents are OpenXML workbooks served through the WordPress document manager;
older links may be BIFF .xls files. The parser therefore tries openpyxl first
and falls back to xlrd.
"""

from __future__ import annotations

from datetime import datetime, timezone
from html import unescape
from io import BytesIO
import re
import time
import unicodedata

import requests

from fed_rankings_common import build_ranking_row, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import set_state

SOURCE = "fis_italy"
COUNTRY = "ITA"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; FenceSpace/1.0; +https://fencespace.app)"
}

BASE_URL = "https://federscherma.it"
DOCUMENT_SEARCH_URL = f"{BASE_URL}/wp-json/wp/v2/search"
DOCUMENT_POST_URL = f"{BASE_URL}/wp-json/wp/v2/documento"
DOWNLOAD_URL = (
    f"{BASE_URL}/wp-content/plugins/if_document_manager/forceDownload.php?ID_file={{document_id}}"
)

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

CATEGORY_SEARCH = {
    "Senior": {
        "query": "ranking assoluti",
        "required": ("ranking", "assoluti"),
        "excluded": ("master", "gpg", "u23", "cadetti", "paralimpico", "non vedenti"),
    },
    "Junior": {
        "query": "ranking giovani",
        "required": ("ranking", "giovani"),
        "excluded": ("gpg", "cadetti", "master", "paralimpico", "non vedenti"),
    },
}

WEAPON_SHEET_CODES = {"Foil": "F", "Epee": "SP", "Sabre": "SC"}
GENDER_SHEET_CODES = {"Men": "M", "Women": "F"}
CATEGORY_SHEET_CODES = {"Senior": "A", "Junior": "G"}
WEAPON_TITLE_WORDS = {"Foil": "fioretto", "Epee": "spada", "Sabre": "sciabola"}
GENDER_TITLE_WORDS = {"Men": "maschile", "Women": "femminile"}
CATEGORY_TITLE_WORDS = {
    "Senior": ("assoluto", "assoluti"),
    "Junior": ("giovani", "junior"),
}

_RANK_HEADERS = {"pos", "posizione", "rank", "#"}
_NAME_HEADERS = {"atleta", "nome", "athlete", "name"}
_CLUB_HEADERS = {"societa", "club", "societa sportiva", "affiliazione"}
_POINTS_HEADERS = {"punti", "points", "punteggio", "pt", "pts", "totale", "total"}


def _strip_accents(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalise_header(value) -> str:
    text = _strip_accents(str(value or "").lower().strip())
    text = re.sub(r"[^a-z0-9#]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _normalise_text(value: str) -> str:
    text = _strip_accents(value.lower())
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_rank(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    match = re.match(r"^\s*(\d+)", str(value).strip())
    if not match:
        return None
    return int(match.group(1))


def _parse_points(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip().replace("\xa0", "").replace(" ", "")
    if not text:
        return None

    comma_pos = text.rfind(",")
    dot_pos = text.rfind(".")
    if comma_pos != -1 and dot_pos != -1:
        if comma_pos > dot_pos:
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif comma_pos != -1:
        if re.match(r"^\d{1,3}(,\d{3})+$", text):
            text = text.replace(",", "")
        else:
            text = text.replace(",", ".")
    elif dot_pos != -1 and re.match(r"^\d{1,3}(\.\d{3})+$", text):
        text = text.replace(".", "")

    try:
        return float(text)
    except ValueError:
        return None


def _detect_columns(row: list) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, value in enumerate(row):
        key = _normalise_header(value)
        if key in _RANK_HEADERS and "rank_col" not in mapping:
            mapping["rank_col"] = idx
        elif key in _NAME_HEADERS and "name_col" not in mapping:
            mapping["name_col"] = idx
        elif key in _CLUB_HEADERS and "club_col" not in mapping:
            mapping["club_col"] = idx
        elif key in _POINTS_HEADERS and "points_col" not in mapping:
            mapping["points_col"] = idx
    return mapping


def _sheet_matches_combo(
    sheet_name: str,
    rows: list[list],
    *,
    weapon: str | None = None,
    gender: str | None = None,
    category: str | None = None,
) -> bool:
    if not any((weapon, gender, category)):
        return True

    if weapon and gender and category:
        expected = (
            f"{WEAPON_SHEET_CODES.get(weapon, '')}"
            f"{GENDER_SHEET_CODES.get(gender, '')}"
            f"{CATEGORY_SHEET_CODES.get(category, '')}"
        )
        compact_name = re.sub(r"[^A-Z0-9]", "", sheet_name.upper())
        if expected and compact_name == expected:
            return True

    sample_cells = [sheet_name]
    for row in rows[:6]:
        sample_cells.extend(_cell_text(value) for value in row[:8])
    text = _normalise_text(" ".join(sample_cells))

    if weapon and WEAPON_TITLE_WORDS.get(weapon, "") not in text:
        return False
    if gender and GENDER_TITLE_WORDS.get(gender, "") not in text:
        return False
    if category and not any(word in text for word in CATEGORY_TITLE_WORDS.get(category, ())):
        return False
    return True


def _parse_sheet_rows(rows: list[list]) -> list[dict]:
    header_map: dict[str, int] | None = None
    parsed: list[dict] = []

    for row in rows:
        values = list(row)
        if header_map is None:
            candidate = _detect_columns(values)
            if "rank_col" in candidate and "name_col" in candidate:
                header_map = candidate
            continue

        max_col = max(header_map.values())
        if len(values) <= max_col:
            continue

        rank = _parse_rank(values[header_map["rank_col"]])
        name = _cell_text(values[header_map["name_col"]])
        if rank is None or not name:
            continue

        club = None
        if "club_col" in header_map:
            club = _cell_text(values[header_map["club_col"]]) or None

        points = None
        if "points_col" in header_map:
            points = _parse_points(values[header_map["points_col"]])

        parsed.append(
            {
                "rank": rank,
                "name": name,
                "club": club,
                "points": points,
            }
        )

    return parsed


def _parse_openpyxl_workbook(
    file_bytes: bytes,
    *,
    weapon: str | None = None,
    gender: str | None = None,
    category: str | None = None,
) -> list[dict]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        results: list[dict] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            if _sheet_matches_combo(sheet.title, rows, weapon=weapon, gender=gender, category=category):
                results.extend(_parse_sheet_rows(rows))
        return results
    finally:
        workbook.close()


def _parse_xlrd_workbook(
    file_bytes: bytes,
    *,
    weapon: str | None = None,
    gender: str | None = None,
    category: str | None = None,
) -> list[dict]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=file_bytes)
    results: list[dict] = []
    for sheet in workbook.sheets():
        rows = [sheet.row_values(row_idx) for row_idx in range(sheet.nrows)]
        if _sheet_matches_combo(sheet.name, rows, weapon=weapon, gender=gender, category=category):
            results.extend(_parse_sheet_rows(rows))
    return results


def parse_rankings_xls(
    file_bytes: bytes,
    *,
    weapon: str | None = None,
    gender: str | None = None,
    category: str | None = None,
) -> list[dict]:
    """Parse ranking rows from an XLSX workbook, falling back to BIFF XLS."""
    if not file_bytes:
        return []

    try:
        return _parse_openpyxl_workbook(
            file_bytes, weapon=weapon, gender=gender, category=category
        )
    except Exception:
        pass

    try:
        return _parse_xlrd_workbook(file_bytes, weapon=weapon, gender=gender, category=category)
    except Exception as exc:
        raise ValueError(
            "Could not parse Federscherma ranking workbook with openpyxl or xlrd"
        ) from exc


def _search_documents(query: str) -> list[dict]:
    response = requests.get(
        DOCUMENT_SEARCH_URL,
        headers=HEADERS,
        params={"search": query, "per_page": 30},
        timeout=20,
    )
    response.raise_for_status()
    data = response.json()
    return data if isinstance(data, list) else []


def _document_by_slug(slug: str) -> dict | None:
    response = requests.get(
        DOCUMENT_POST_URL,
        headers=HEADERS,
        params={"slug": slug},
        timeout=20,
    )
    response.raise_for_status()
    data = response.json()
    if isinstance(data, list) and data:
        return data[0]
    return None


def _slug_from_url(url: str) -> str | None:
    clean = url.rstrip("/")
    if not clean:
        return None
    return clean.split("/")[-1] or None


def _title_matches(title: str, *, required: tuple[str, ...], excluded: tuple[str, ...]) -> bool:
    text = _normalise_text(unescape(title))
    return all(word in text for word in required) and not any(word in text for word in excluded)


def discover_latest_ranking_documents() -> dict[str, dict]:
    """Return latest Federscherma ranking document metadata for Senior and Junior."""
    documents: dict[str, dict] = {}

    for category, criteria in CATEGORY_SEARCH.items():
        for item in _search_documents(str(criteria["query"])):
            title = unescape(str(item.get("title") or ""))
            if not _title_matches(
                title,
                required=tuple(criteria["required"]),
                excluded=tuple(criteria["excluded"]),
            ):
                continue

            page_url = str(item.get("url") or item.get("link") or "")
            slug = _slug_from_url(page_url)
            document = _document_by_slug(slug) if slug else None
            document_id = document.get("id") if document else item.get("id")
            if not document_id:
                continue

            documents[category] = {
                "category": category,
                "title": title,
                "page_url": page_url,
                "document_id": int(document_id),
                "download_url": DOWNLOAD_URL.format(document_id=int(document_id)),
            }
            break

    return documents


def download_ranking_file(url: str) -> bytes:
    response = requests.get(url, headers=HEADERS, timeout=40)
    response.raise_for_status()
    return response.content


def current_season() -> str:
    now = datetime.now(timezone.utc)
    year = now.year
    return f"{year-1}-{year}" if now.month < 7 else f"{year}-{year+1}"


def _combo_label(weapon: str, gender: str, category: str) -> str:
    return f"{category} {gender} {weapon}"


def main():
    run_log = ScraperRunLogger("scrape_fed_italy").start()
    season = current_season()
    print(f"FIS Italy rankings — season {season}")

    total_written = 0
    total_failed = 0
    total_skipped = 0
    combos_with_rows: list[str] = []
    file_cache: dict[str, bytes] = {}

    try:
        documents = discover_latest_ranking_documents()
        if documents:
            set_state(SOURCE, "latest_documents", documents)
        else:
            print("  No Federscherma ranking documents found")

        for category, document in documents.items():
            print(f"  {category} document: {document['title']} -> {document['download_url']}")

        for weapon, gender, category in RANKING_COMBOS:
            label = _combo_label(weapon, gender, category)
            print(f"  {label}...")
            ranking_document = documents.get(category)
            if not ranking_document:
                print("    No ranking workbook found")
                total_skipped += 1
                continue

            try:
                if category not in file_cache:
                    file_cache[category] = download_ranking_file(ranking_document["download_url"])
                parsed = parse_rankings_xls(
                    file_cache[category],
                    weapon=weapon,
                    gender=gender,
                    category=category,
                )
            except Exception as exc:
                print(f"    Failed to download/parse workbook: {exc}")
                total_failed += 1
                time.sleep(REQUEST_DELAY)
                continue

            if not parsed:
                print("    No rows parsed")
                total_skipped += 1
                time.sleep(REQUEST_DELAY)
                continue

            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                    metadata={
                        "source_page": document["page_url"],
                        "download_url": document["download_url"],
                    },
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            print(f"    Parsed {len(rows)} rows; written {written} rows")
            combos_with_rows.append(label)
            total_written += written
            time.sleep(REQUEST_DELAY)

        metadata = {
            "documents": documents,
            "combos_with_rows": combos_with_rows,
            "combos_attempted": len(RANKING_COMBOS),
        }
        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata=metadata,
        )
        print(
            f"Done — written={total_written}, failed={total_failed}, skipped={total_skipped}, "
            f"combos_with_rows={len(combos_with_rows)}/{len(RANKING_COMBOS)}"
        )
    except Exception as exc:
        run_log.error(str(exc))
        print(f"Fatal error: {exc}")


if __name__ == "__main__":
    main()
