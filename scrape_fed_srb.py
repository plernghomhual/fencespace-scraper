"""
scrape_fed_srb.py - Serbia national federation rankings scraper.

Probe evidence, 2026-06-02:
  - Supplied probe host `macesavez.rs` failed DNS resolution locally.
  - Active official public site: https://www.mss.org.rs/
  - Working ranking URL: https://www.mss.org.rs/rang-liste/
  - Request method: GET
  - Response format: WordPress HTML page with a public MSS ranking download.
    The download can be HTML/PDF/Excel depending on the published package.
  - Public page advertises "Rang Liste MSS (26.04.2026)" for all categories
    and disciplines; combo-level availability is determined from the extracted
    ranking text at runtime.
"""

from __future__ import annotations

import html
import io
import re
import time
import unicodedata
from datetime import UTC, datetime, timezone
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

from fed_rankings_common import build_ranking_row, federation_request, write_rankings
from run_logger import ScraperRunLogger
from scraper_state import set_state

try:
    from season_utils import season_to_string
except ImportError:  # pragma: no cover - compatibility fallback
    def season_to_string(season_int: int) -> str:
        return f"{season_int - 1:04d}-{season_int:04d}"


SOURCE = "srb_fencing"
COUNTRY = "Serbia"
BASE_URL = "https://www.mss.org.rs/rang-liste/"
LEGACY_PROBE_URL = "https://macesavez.rs/"
REQUEST_DELAY = 1.5
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "sr,en;q=0.9",
}

RANKING_COMBOS = [
    ("Foil", "Men", "Senior"),
    ("Foil", "Women", "Senior"),
    ("Epee", "Men", "Senior"),
    ("Epee", "Women", "Senior"),
    ("Sabre", "Men", "Senior"),
    ("Sabre", "Women", "Senior"),
    ("Foil", "Men", "Junior"),
    ("Foil", "Women", "Junior"),
    ("Epee", "Men", "Junior"),
    ("Epee", "Women", "Junior"),
    ("Sabre", "Men", "Junior"),
    ("Sabre", "Women", "Junior"),
]

_RANK_HEADERS = {"plasman", "pozicija", "rank", "poredak", "пласман", "позиција"}
_NAME_HEADERS = {
    "imeiprezime",
    "takmicar",
    "takmicarka",
    "ime",
    "name",
    "fencer",
    "имеипрезиме",
    "такмичар",
    "такмичарка",
    "име",
}
_CLUB_HEADERS = {"klub", "club", "клуб"}
_POINT_HEADERS = {"bodovi", "poeni", "points", "pts", "total", "бодови", "поени"}
_SKIP_VALUES = {
    "dns",
    "dnf",
    "dq",
    "dsq",
    "wd",
    "ret",
    "total",
    "summary",
    "ukupno",
    "zbir",
    "резиме",
    "укупно",
    "збир",
}
_DIRECT_DOWNLOAD_MARKERS = (".xlsx", ".xls", ".pdf", "wpdmdl=", "/download/")
_RANKING_TEXT_CACHE: str | None = None
_RANKING_SOURCE_URL: str | None = None
_RANKING_FAILURE_REASON: str | None = None

_CYRILLIC_TO_LATIN = str.maketrans(
    {
        "а": "a",
        "б": "b",
        "в": "v",
        "г": "g",
        "д": "d",
        "ђ": "dj",
        "е": "e",
        "ж": "z",
        "з": "z",
        "и": "i",
        "ј": "j",
        "к": "k",
        "л": "l",
        "љ": "lj",
        "м": "m",
        "н": "n",
        "њ": "nj",
        "о": "o",
        "п": "p",
        "р": "r",
        "с": "s",
        "т": "t",
        "ћ": "c",
        "у": "u",
        "ф": "f",
        "х": "h",
        "ц": "c",
        "ч": "c",
        "џ": "dz",
        "ш": "s",
    }
)

_WEAPON_TOKENS = {
    "Foil": {"foil", "floret", "флорет"},
    "Epee": {"epee", "épée", "mac", "mač", "мач", "spada"},
    "Sabre": {"sabre", "saber", "sablja", "сабља"},
}
_GENDER_TOKENS = {
    "Men": {"men", "muskarci", "muski", "muški", "мушки", "мушкарци", "seniori", "juniori"},
    "Women": {"women", "zene", "žene", "zenski", "ženski", "жене", "женски", "seniorke", "juniorke"},
}
_CATEGORY_TOKENS = {
    "Senior": {"senior", "seniori", "seniorke", "сениори", "сениорке"},
    "Junior": {"junior", "juniori", "juniorke", "u20", "јуниори", "јуниорке"},
}
_EXPLICIT_MEN_TOKENS = {"men", "muskarci", "muski", "muški", "мушки", "мушкарци"}
_EXPLICIT_WOMEN_TOKENS = {
    "women",
    "zene",
    "žene",
    "zenski",
    "ženski",
    "жене",
    "женски",
    "seniorke",
    "juniorke",
}


def current_season() -> str:
    """Return the current competition season as YYYY-YYYY."""
    now = datetime.now(UTC)
    season_end_year = now.year if now.month < 7 else now.year + 1
    return season_to_string(season_end_year)


def _strip_accents(value: str) -> str:
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", value) if not unicodedata.combining(ch)
    )


def _compact_text(value: str | None) -> str:
    return re.sub(r"\s+", " ", value or "").strip()


def _header_key(value: str) -> str:
    value = _strip_accents(_compact_text(value).lower())
    return re.sub(r"[^0-9a-zа-яђјљњћџшчж]+", "", value)


def _search_key(value: str) -> str:
    value = _strip_accents(_compact_text(value).lower()).translate(_CYRILLIC_TO_LATIN)
    return re.sub(r"[^0-9a-z]+", " ", value).strip()


def _parse_rank(value: str) -> int | None:
    match = re.match(r"\s*(\d+)", _compact_text(value))
    if not match:
        return None
    rank = int(match.group(1))
    return rank if rank > 0 else None


def _parse_points(value: str) -> float | None:
    raw = _compact_text(value)
    if not raw:
        return None
    cleaned = re.sub(r"[^0-9,.\-]", "", raw)
    if "," in cleaned and "." in cleaned and cleaned.rfind(",") > cleaned.rfind("."):
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
    if not match:
        return None
    return float(match.group(0))


def _is_skip_row(values: list[str]) -> bool:
    keys = {_header_key(value) for value in values if value}
    return any(key in _SKIP_VALUES for key in keys)


def _header_mapping(values: list[str]) -> dict[str, int] | None:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values):
        key = _header_key(value)
        if key in _RANK_HEADERS:
            mapping["rank"] = index
        elif key in _NAME_HEADERS:
            mapping["name"] = index
        elif key in _CLUB_HEADERS:
            mapping["club"] = index
        elif key in _POINT_HEADERS:
            mapping["points"] = index
    if {"rank", "name", "points"}.issubset(mapping):
        return mapping
    return None


def _row_from_values(values: list[str], mapping: dict[str, int]) -> dict | None:
    if _is_skip_row(values):
        return None
    max_index = max(mapping.values())
    if len(values) <= max_index:
        return None

    rank = _parse_rank(values[mapping["rank"]])
    name = _compact_text(values[mapping["name"]])
    points = _parse_points(values[mapping["points"]])
    if rank is None or not name or points is None:
        return None
    if _header_key(name) in _SKIP_VALUES:
        return None

    club = None
    if "club" in mapping and mapping["club"] < len(values):
        club = _compact_text(values[mapping["club"]]) or None
    return {"rank": rank, "name": name, "club": club, "points": points}


def _split_text_row(line: str) -> list[str]:
    line = _compact_text(line)
    if not line:
        return []
    if "|" in line:
        return [part.strip() for part in line.split("|")]
    if "\t" in line:
        return [part.strip() for part in line.split("\t")]
    parts = [part.strip() for part in re.split(r"\s{2,}", line) if part.strip()]
    if len(parts) >= 4:
        return parts
    match = re.match(
        r"^(\d+)[.)]?\s+(.+?)\s{2,}(.+?)\s+([0-9][0-9,.\s]*)$",
        line,
    )
    if match:
        return [match.group(1), match.group(2), match.group(3), match.group(4)]
    return [line]


def _parse_html_tables(html_or_text: str) -> list[dict]:
    soup = BeautifulSoup(html_or_text, "html.parser")
    parsed: list[dict] = []
    for table in soup.find_all("table"):
        mapping: dict[str, int] | None = None
        for tr in table.find_all("tr"):
            cells = tr.find_all(["th", "td"])
            values = [_compact_text(cell.get_text(" ", strip=True)) for cell in cells]
            if not values:
                continue
            maybe_header = _header_mapping(values)
            if maybe_header:
                mapping = maybe_header
                continue
            if mapping:
                row = _row_from_values(values, mapping)
                if row:
                    parsed.append(row)
    return parsed


def _parse_text_rows(html_or_text: str) -> list[dict]:
    soup = BeautifulSoup(html_or_text, "html.parser")
    text = soup.get_text("\n") if soup.find() else html_or_text
    parsed: list[dict] = []
    mapping: dict[str, int] | None = None
    for line in text.splitlines():
        values = _split_text_row(line)
        if not values:
            continue
        maybe_header = _header_mapping(values)
        if maybe_header:
            mapping = maybe_header
            continue
        if mapping:
            row = _row_from_values(values, mapping)
            if row:
                parsed.append(row)
    return parsed


def parse_rankings_table(html_or_text: str) -> list[dict]:
    """Parse Serbian rankings from HTML, PDF-extracted text, or workbook text."""
    if not html_or_text or not html_or_text.strip():
        return []

    html_rows = _parse_html_tables(html_or_text)
    if html_rows:
        return html_rows
    return _parse_text_rows(html_or_text)


def _page_is_blocked_or_login_only(status_code: int, body: str) -> bool:
    if status_code in {401, 403, 429}:
        return True
    lowered = body.lower()
    has_public_ranking_marker = any(marker in lowered for marker in ("rang", "preuzmi", "wpdm"))
    if "captcha" in lowered or "access denied" in lowered or "cloudflare" in lowered:
        return True
    if ("loginform" in lowered or "type=\"password\"" in lowered or "wp-login" in lowered) and not has_public_ranking_marker:
        return True
    if "please enable javascript" in lowered and not has_public_ranking_marker:
        return True
    return False


def _get_url(url: str):
    global _RANKING_FAILURE_REASON
    try:
        response = federation_request(
            "GET",
            url,
            headers=HEADERS,
            timeout=20,
            allow_redirects=True,
        )
    except requests.RequestException as exc:
        _RANKING_FAILURE_REASON = f"network error for {url}: {exc}"
        print(f"    Request error for {url}: {exc}")
        return None
    except Exception as exc:
        _RANKING_FAILURE_REASON = f"request failed for {url}: {exc}"
        print(f"    Request error for {url}: {exc}")
        return None

    status_code = int(getattr(response, "status_code", 0) or 0)
    body = getattr(response, "text", "") or ""
    if status_code != 200:
        _RANKING_FAILURE_REASON = f"HTTP {status_code} for {url}"
        print(f"    HTTP {status_code} for {url}")
        return None
    if _page_is_blocked_or_login_only(status_code, body):
        _RANKING_FAILURE_REASON = f"blocked/login/js-only page at {url}"
        print(f"    Blocked, login-only, or JS-only page at {url}")
        return None
    return response


def _content_type(response) -> str:
    headers = getattr(response, "headers", {}) or {}
    return str(headers.get("content-type") or headers.get("Content-Type") or "").lower()


def _response_bytes(response) -> bytes:
    content = getattr(response, "content", None)
    if isinstance(content, bytes):
        return content
    return (getattr(response, "text", "") or "").encode("utf-8", errors="ignore")


def _excel_to_text(file_bytes: bytes) -> str | None:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"SHEET: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [_compact_text(str(cell)) for cell in row if cell not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
            lines.append("")
        return "\n".join(lines)
    except Exception:
        pass

    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=file_bytes)
        lines = []
        for sheet in workbook.sheets():
            lines.append(f"SHEET: {sheet.name}")
            for row_index in range(sheet.nrows):
                values = [
                    _compact_text(str(sheet.cell_value(row_index, col_index)))
                    for col_index in range(sheet.ncols)
                    if sheet.cell_value(row_index, col_index) not in (None, "")
                ]
                if values:
                    lines.append(" | ".join(values))
            lines.append("")
        return "\n".join(lines)
    except Exception as exc:
        print(f"    Could not parse ranking workbook: {exc}")
        return None


def _pdf_to_text(file_bytes: bytes) -> str | None:
    try:
        import pdfplumber

        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            return "\n".join(
                page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages
            )
    except Exception as exc:
        print(f"    Could not parse ranking PDF: {exc}")
        return None


def _response_to_rankings_text(response, url: str) -> str | None:
    body = getattr(response, "text", "") or ""
    content = _response_bytes(response)
    content_type = _content_type(response)
    lowered_url = url.lower()

    if ".xlsx" in lowered_url or "spreadsheet" in content_type or content.startswith(b"PK\x03\x04"):
        return _excel_to_text(content)
    if ".xls" in lowered_url or "excel" in content_type:
        return _excel_to_text(content)
    if ".pdf" in lowered_url or "pdf" in content_type or content.startswith(b"%PDF"):
        return _pdf_to_text(content)
    return body


def _find_download_candidates(page_html: str, base_url: str) -> list[str]:
    candidates: list[str] = []

    def add(raw_url: str | None) -> None:
        if not raw_url:
            return
        raw_url = html.unescape(str(raw_url).strip().strip("\"'"))
        if not raw_url or raw_url.startswith(("mailto:", "tel:", "#")):
            return
        absolute = urljoin(base_url, raw_url)
        lowered = absolute.lower()
        if not any(marker in lowered for marker in _DIRECT_DOWNLOAD_MARKERS):
            return
        if absolute not in candidates:
            candidates.append(absolute)

    soup = BeautifulSoup(page_html, "html.parser")
    for tag in soup.find_all(True):
        add(tag.get("href"))
        add(tag.get("src"))
        add(tag.get("action"))
        for attr, value in tag.attrs.items():
            if attr.startswith("data-") or attr in {"onclick", "value"}:
                if isinstance(value, list):
                    for item in value:
                        add(str(item))
                else:
                    add(str(value))

    for match in re.finditer(
        r"""(?:https?://[^\s"'<>]+|/[^\s"'<>]*(?:download|wpdmdl=)[^\s"'<>]*|\?wpdmdl=\d+[^\s"'<>]*)""",
        page_html,
        re.IGNORECASE,
    ):
        add(match.group(0))

    return candidates


def _download_from_candidate(url: str, depth: int = 0) -> str | None:
    global _RANKING_SOURCE_URL
    if depth > 2:
        return None
    response = _get_url(url)
    if response is None:
        return None

    text = _response_to_rankings_text(response, getattr(response, "url", url))
    if text and parse_rankings_table(text):
        _RANKING_SOURCE_URL = getattr(response, "url", url)
        return text

    content_type = _content_type(response)
    if "html" not in content_type:
        return text

    for candidate in _find_download_candidates(text or "", getattr(response, "url", url)):
        if candidate == url:
            continue
        downloaded = _download_from_candidate(candidate, depth + 1)
        if downloaded:
            return downloaded
    return text


def _download_latest_ranking_text() -> str | None:
    global _RANKING_TEXT_CACHE, _RANKING_SOURCE_URL, _RANKING_FAILURE_REASON
    if _RANKING_TEXT_CACHE is not None:
        return _RANKING_TEXT_CACHE or None

    response = _get_url(BASE_URL)
    if response is None:
        _RANKING_TEXT_CACHE = ""
        return None

    page_url = getattr(response, "url", BASE_URL)
    page_html = getattr(response, "text", "") or ""
    for candidate in _find_download_candidates(page_html, page_url):
        downloaded = _download_from_candidate(candidate)
        if downloaded:
            _RANKING_TEXT_CACHE = downloaded
            return downloaded

    _RANKING_SOURCE_URL = page_url
    _RANKING_TEXT_CACHE = page_html
    if not parse_rankings_table(page_html):
        _RANKING_FAILURE_REASON = f"No scrapeable rankings at {BASE_URL}"
    return page_html


def _combo_token_sets(weapon: str, gender: str, category: str) -> tuple[set[str], set[str], set[str]]:
    return (
        {_search_key(token) for token in _WEAPON_TOKENS.get(weapon, {weapon})},
        {_search_key(token) for token in _GENDER_TOKENS.get(gender, {gender})},
        {_search_key(token) for token in _CATEGORY_TOKENS.get(category, {category})},
    )


def _line_matches_combo(line: str, weapon: str, gender: str, category: str) -> bool:
    haystack = f" {_search_key(line)} "
    weapon_tokens, gender_tokens, category_tokens = _combo_token_sets(weapon, gender, category)
    explicit_men = {_search_key(token) for token in _EXPLICIT_MEN_TOKENS}
    explicit_women = {_search_key(token) for token in _EXPLICIT_WOMEN_TOKENS}
    if gender == "Men" and any(f" {token} " in haystack for token in explicit_women):
        return False
    if gender == "Women" and any(f" {token} " in haystack for token in explicit_men):
        return False
    return (
        any(f" {token} " in haystack for token in weapon_tokens)
        and any(f" {token} " in haystack for token in gender_tokens)
        and any(f" {token} " in haystack for token in category_tokens)
    )


def _line_matches_any_combo(line: str) -> bool:
    return any(_line_matches_combo(line, weapon, gender, category) for weapon, gender, category in RANKING_COMBOS)


def _extract_combo_section(text: str, weapon: str, gender: str, category: str) -> str | None:
    lines = text.splitlines()
    section_starts = [index for index, line in enumerate(lines) if _line_matches_any_combo(line)]
    if not section_starts:
        return text

    for pos, start in enumerate(section_starts):
        if not _line_matches_combo(lines[start], weapon, gender, category):
            continue
        end = section_starts[pos + 1] if pos + 1 < len(section_starts) else len(lines)
        section = "\n".join(lines[start:end]).strip()
        return section or None
    return None


def fetch_rankings_page(weapon: str, gender: str, category: str) -> str | None:
    """Return extracted ranking text for one weapon/gender/category combo."""
    combo = (weapon, gender, category)
    if combo not in RANKING_COMBOS:
        print(f"    Unsupported combo {combo}")
        return None

    text = _download_latest_ranking_text()
    if not text:
        print(f"    No scrapeable rankings at {BASE_URL}")
        return None

    section = _extract_combo_section(text, weapon, gender, category)
    if not section:
        print(f"    No public combo section for {weapon} {gender} {category}")
        return None
    return section


def main() -> None:
    run_log = ScraperRunLogger("scrape_fed_srb").start()
    season = current_season()
    total_written = 0
    total_failed = 0
    total_skipped = 0
    failed_combos: list[str] = []

    print(f"Serbia federation rankings - season {season}")
    print(f"Source URL: {BASE_URL}")
    print(f"Legacy probe URL unavailable locally: {LEGACY_PROBE_URL}")

    try:
        for weapon, gender, category in RANKING_COMBOS:
            print(f"  {weapon} {gender} {category}...")
            content = fetch_rankings_page(weapon, gender, category)
            if not content:
                total_failed += 1
                failed_combos.append(f"{weapon} {gender} {category}: no public data")
                time.sleep(REQUEST_DELAY)
                continue

            parsed = parse_rankings_table(content)
            if not parsed:
                total_failed += 1
                failed_combos.append(f"{weapon} {gender} {category}: no rows parsed")
                print("    No rows parsed")
                time.sleep(REQUEST_DELAY)
                continue

            rows = [
                build_ranking_row(
                    source=SOURCE,
                    season=season,
                    weapon=weapon,
                    gender=gender,
                    category=category,
                    rank=row["rank"],
                    name=row["name"],
                    country=COUNTRY,
                    club=row.get("club"),
                    points=row.get("points"),
                )
                for row in parsed
            ]
            written = write_rankings(rows, source=SOURCE, season=season)
            total_written += written
            print(f"    Parsed {len(parsed)} rows; written {written}")
            time.sleep(REQUEST_DELAY)

        if total_failed == len(RANKING_COMBOS) and total_written == 0:
            print(f"No scrapeable rankings at {BASE_URL}")
            if _RANKING_FAILURE_REASON:
                print(f"Probe evidence: {_RANKING_FAILURE_REASON}")

        set_state(
            SOURCE,
            "last_run",
            {
                "season": season,
                "source_url": _RANKING_SOURCE_URL or BASE_URL,
                "failed_combos": failed_combos,
                "failure_reason": _RANKING_FAILURE_REASON,
                "updated_at": datetime.now(UTC).isoformat(),
            },
        )
        run_log.complete(
            written=total_written,
            failed=total_failed,
            skipped=total_skipped,
            metadata={
                "season": season,
                "source_url": _RANKING_SOURCE_URL or BASE_URL,
                "failed_combos": failed_combos,
                "failure_reason": _RANKING_FAILURE_REASON,
            },
        )
        print(
            f"Done - written={total_written}, failed={total_failed}, skipped={total_skipped}"
        )
    except Exception as exc:
        run_log.error(str(exc))
        raise


if __name__ == "__main__":
    main()
